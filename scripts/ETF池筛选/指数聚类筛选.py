#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按月对ETF对应指数进行收益来源去重，生成每月底动态指数池。

处理逻辑：
1. 读取每月ETF初筛结果及对应指数60个共同交易区间收益率；
2. 检查全部指数的60个收益日期完全一致，计算指数收益率相关矩阵；
3. 将相关矩阵转换为距离矩阵：Distance = 1 - Corr；
4. 使用complete linkage层次聚类，阈值由参数区设定；
5. 每个聚类仅保留过去20个市场交易日平均成交额最大的ETF及其对应指数。

输入：
- outputs/etf_pool/initial/YYYY_MM_DD.csv
- outputs/etf_pool/index_returns/window_60/YYYY_MM_DD.csv
- outputs/etf_data/etf_data.csv（用于计算20日平均成交额）

输出：
- outputs/etf_pool/clusters/threshold_<相关性阈值>/reports/YYYY_MM_DD.xlsx
- 每个工作簿的第一张表为月底动态指数池；
- 第二张表为代表筛选前所有原始指数的完整相关性矩阵。
- 第三张表为每个聚类的全部成员、代表ETF及聚类内相关性。
"""

from __future__ import annotations

import csv
import math
from bisect import bisect_right
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from scipy.cluster.hierarchy import fcluster, linkage
from scipy.spatial.distance import squareform


PROJECT_ROOT = Path(__file__).resolve().parents[2]
INITIAL_SELECTION_DIR = PROJECT_ROOT / "outputs" / "etf_pool" / "initial"
ETF_DATA_FILE = PROJECT_ROOT / "outputs" / "etf_data" / "etf_data.csv"
# ============================= 聚类参数 =============================
# 可改为0.7、0.8或0.9；距离阈值自动按1 - CORRELATION_THRESHOLD计算。
CORRELATION_THRESHOLD = 0.9
ALLOWED_CORRELATION_THRESHOLDS = (0.7, 0.8, 0.9)

RETURN_TRADING_DAYS = 60
TURNOVER_LOOKBACK_DAYS = 20
# ====================================================================

INDEX_RETURN_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_pool"
    / "index_returns"
    / f"window_{RETURN_TRADING_DAYS}"
)
OUTPUT_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_pool"
    / "clusters"
    / f"threshold_{CORRELATION_THRESHOLD:g}"
    / "reports"
)

INITIAL_REQUIRED_COLUMNS = {
    "日期",
    "代码",
    "对标指数",
    "对标指数代码",
}
INDEX_REQUIRED_COLUMNS = {
    "月末交易日",
    "收益日期",
    "对标指数代码",
    "日收益率",
}
ETF_DATA_REQUIRED_COLUMNS = {"日期", "代码", "成交额"}
ADDED_OUTPUT_COLUMNS = [
    "过去20个交易日平均成交额",
    "聚类编号",
    "聚类指数数量",
]
CLUSTER_DETAIL_COLUMNS = [
    "聚类编号",
    "是否最终代表",
    "ETF代码",
    "ETF名称",
    "类别大类",
    "对标指数代码",
    "对标指数",
    "过去20个交易日平均成交额",
    "与代表指数相关性",
    "聚类内最低相关性",
    "相关性阈值",
    "距离阈值",
]
@dataclass(frozen=True)
class MonthInput:
    """一个月末的ETF初筛数据和指数收益率文件。"""

    selection_date: date
    initial_file: Path
    index_return_file: Path
    etf_rows: tuple[Mapping[str, str], ...]

    @property
    def file_name(self) -> str:
        return self.selection_date.strftime("%Y_%m_%d.xlsx")

    @property
    def etf_codes(self) -> set[str]:
        return {clean_text(row.get("代码")) for row in self.etf_rows}

    @property
    def index_codes(self) -> set[str]:
        return {clean_text(row.get("对标指数代码")) for row in self.etf_rows}

def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "--", "None", "null", "NULL"} else text


def parse_date(value: object, field_name: str) -> date:
    text = clean_text(value)
    # 大部分数据为YYYY-MM-DD，先走快速路径。
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        try:
            return date(int(text[:4]), int(text[5:7]), int(text[8:10]))
        except ValueError:
            pass
    for date_format in ("%Y%m%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name}不是有效日期：{text!r}")


def parse_finite_number(value: object) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def read_initial_month(path: Path) -> tuple[list[str], tuple[Mapping[str, str], ...]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        missing_columns = INITIAL_REQUIRED_COLUMNS - set(fieldnames)
        if missing_columns:
            raise ValueError(f"{path.name}缺少列：{sorted(missing_columns)}")
        rows = tuple(dict(row) for row in reader)

    for row_number, row in enumerate(rows, start=2):
        if not clean_text(row.get("代码")):
            raise ValueError(f"{path.name}第{row_number}行ETF代码为空")
        if not clean_text(row.get("对标指数代码")):
            raise ValueError(f"{path.name}第{row_number}行对标指数代码为空")
    return fieldnames, rows


def discover_month_inputs() -> tuple[list[str], list[MonthInput]]:
    if not INITIAL_SELECTION_DIR.exists():
        raise FileNotFoundError(f"找不到ETF初筛目录：{INITIAL_SELECTION_DIR}")
    if not INDEX_RETURN_DIR.exists():
        raise FileNotFoundError(f"找不到指数收益率目录：{INDEX_RETURN_DIR}")

    initial_files = sorted(INITIAL_SELECTION_DIR.glob("*.csv"))
    if not initial_files:
        raise FileNotFoundError(f"ETF初筛目录没有CSV：{INITIAL_SELECTION_DIR}")

    common_fieldnames: list[str] | None = None
    months: list[MonthInput] = []
    for initial_file in initial_files:
        try:
            file_date = datetime.strptime(initial_file.stem, "%Y_%m_%d").date()
        except ValueError as exc:
            raise ValueError(f"ETF初筛文件名日期无效：{initial_file.name}") from exc

        index_return_file = INDEX_RETURN_DIR / initial_file.name
        if not index_return_file.exists():
            raise FileNotFoundError(
                f"{initial_file.name}缺少对应指数收益率文件："
                f"{index_return_file}"
            )

        fieldnames, rows = read_initial_month(initial_file)
        if common_fieldnames is None:
            common_fieldnames = fieldnames
        elif fieldnames != common_fieldnames:
            raise ValueError(f"{initial_file.name}的列与其他ETF初筛文件不一致")

        selection_dates = {
            parse_date(row.get("日期"), "日期") for row in rows
        }
        if selection_dates and selection_dates != {file_date}:
            raise ValueError(
                f"{initial_file.name}内的日期与文件名不一致："
                f"{sorted(selection_dates)}"
            )
        months.append(
            MonthInput(
                selection_date=file_date,
                initial_file=initial_file,
                index_return_file=index_return_file,
                etf_rows=rows,
            )
        )

    return common_fieldnames or [], months


def inspect_etf_data_dates(path: Path) -> list[date]:
    if not path.exists():
        raise FileNotFoundError(f"找不到ETF汇总数据：{path}")

    available_dates: set[date] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing_columns = ETF_DATA_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing_columns:
            raise ValueError(f"etf_data.csv缺少列：{sorted(missing_columns)}")
        for row in reader:
            available_dates.add(parse_date(row.get("日期"), "日期"))

    if not available_dates:
        raise ValueError(f"ETF汇总数据中没有有效日期：{path}")
    return sorted(available_dates)


def build_turnover_windows(
    months: Sequence[MonthInput],
    available_dates: Sequence[date],
) -> dict[str, tuple[date, ...]]:
    windows: dict[str, tuple[date, ...]] = {}
    for month in months:
        position = bisect_right(available_dates, month.selection_date)
        window = tuple(
            available_dates[
                position - TURNOVER_LOOKBACK_DAYS : position
            ]
        )
        if len(window) != TURNOVER_LOOKBACK_DAYS:
            raise ValueError(
                f"{month.selection_date}之前不足"
                f"{TURNOVER_LOOKBACK_DAYS}个市场交易日"
            )
        if window[-1] != month.selection_date:
            raise ValueError(f"{month.selection_date}不在etf_data.csv交易日中")
        windows[month.file_name] = window
    return windows


def calculate_average_turnover(
    path: Path,
    months: Sequence[MonthInput],
    windows: Mapping[str, Sequence[date]],
) -> dict[tuple[str, str], float]:
    """按ETF初筛口径计算20个市场交易日平均成交额。

    ETF在某日没有记录或成交额为空时按0计，分母固定为20。
    """

    etf_codes_by_month = {
        month.file_name: month.etf_codes for month in months
    }
    months_by_turnover_date: dict[date, list[str]] = defaultdict(list)
    for file_name, turnover_dates in windows.items():
        for turnover_date in turnover_dates:
            months_by_turnover_date[turnover_date].append(file_name)

    # 保留到“月份 + ETF + 日期”级别，避免源文件意外重复行被重复累加。
    amounts: dict[str, dict[str, dict[date, float]]] = {
        month.file_name: defaultdict(dict) for month in months
    }
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            current_date = parse_date(row.get("日期"), "日期")
            related_months = months_by_turnover_date.get(current_date)
            if not related_months:
                continue
            etf_code = clean_text(row.get("代码"))
            amount = parse_finite_number(row.get("成交额")) or 0.0
            for file_name in related_months:
                if etf_code in etf_codes_by_month[file_name]:
                    amounts[file_name][etf_code][current_date] = amount

    averages: dict[tuple[str, str], float] = {}
    for month in months:
        for etf_code in month.etf_codes:
            total_amount = sum(amounts[month.file_name][etf_code].values())
            averages[(month.file_name, etf_code)] = (
                total_amount / TURNOVER_LOOKBACK_DAYS
            )
    return averages


def read_month_index_returns(
    month: MonthInput,
) -> dict[str, dict[date, float]]:
    returns_by_code: dict[str, dict[date, float]] = defaultdict(dict)
    with month.index_return_file.open(
        "r", encoding="utf-8-sig", newline=""
    ) as handle:
        reader = csv.DictReader(handle)
        missing_columns = INDEX_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing_columns:
            raise ValueError(
                f"{month.index_return_file.name}缺少列：{sorted(missing_columns)}"
            )
        for row_number, row in enumerate(reader, start=2):
            month_end = parse_date(row.get("月末交易日"), "月末交易日")
            if month_end != month.selection_date:
                raise ValueError(
                    f"{month.index_return_file.name}第{row_number}行月末交易日"
                    f"与文件名不一致"
                )
            index_code = clean_text(row.get("对标指数代码"))
            if index_code not in month.index_codes:
                continue
            return_date = parse_date(row.get("收益日期"), "收益日期")
            daily_return = parse_finite_number(row.get("日收益率"))
            if daily_return is None:
                raise ValueError(
                    f"{month.index_return_file.name}第{row_number}行日收益率无效"
                )
            if return_date in returns_by_code[index_code]:
                raise ValueError(
                    f"{month.index_return_file.name}存在重复收益率："
                    f"{index_code} {return_date}"
                )
            returns_by_code[index_code][return_date] = daily_return

    missing_codes = sorted(month.index_codes - set(returns_by_code))
    if missing_codes:
        raise ValueError(
            f"{month.index_return_file.name}缺少指数收益率："
            f"{'、'.join(missing_codes)}"
        )
    wrong_length_codes = sorted(
        code
        for code in month.index_codes
        if len(returns_by_code[code]) != RETURN_TRADING_DAYS
    )
    if wrong_length_codes:
        details = "、".join(
            f"{code}({len(returns_by_code[code])}日)"
            for code in wrong_length_codes
        )
        raise ValueError(
            f"{month.index_return_file.name}不是每个指数都有"
            f"{RETURN_TRADING_DAYS}日收益率：{details}"
        )

    ordered_codes = sorted(month.index_codes)
    reference_dates = tuple(sorted(returns_by_code[ordered_codes[0]]))
    mismatched_codes = [
        index_code
        for index_code in ordered_codes[1:]
        if tuple(sorted(returns_by_code[index_code])) != reference_dates
    ]
    if mismatched_codes:
        raise ValueError(
            f"{month.index_return_file.name}的指数收益日期不统一："
            f"{'、'.join(mismatched_codes)}；请先重新运行指数收益率准备.py"
        )
    return dict(returns_by_code)


def calculate_correlation_and_distance_matrices(
    index_codes: Sequence[str],
    returns_by_code: Mapping[str, Mapping[date, float]],
) -> tuple[np.ndarray, np.ndarray]:
    """按全体指数统一的60个收益区间计算相关与距离矩阵。"""

    count = len(index_codes)
    correlations = np.eye(count, dtype=float)
    distances = np.zeros((count, count), dtype=float)
    return_dates = sorted(returns_by_code[index_codes[0]])
    if len(return_dates) != RETURN_TRADING_DAYS:
        raise ValueError(
            f"统一收益日期应为{RETURN_TRADING_DAYS}个，"
            f"实际为{len(return_dates)}个"
        )

    for left_position in range(count):
        left_code = index_codes[left_position]
        left_returns = returns_by_code[left_code]
        for right_position in range(left_position + 1, count):
            right_code = index_codes[right_position]
            right_returns = returns_by_code[right_code]
            left_values = np.asarray(
                [left_returns[current_date] for current_date in return_dates],
                dtype=float,
            )
            right_values = np.asarray(
                [right_returns[current_date] for current_date in return_dates],
                dtype=float,
            )
            left_centered = left_values - left_values.mean()
            right_centered = right_values - right_values.mean()
            denominator = math.sqrt(
                float(np.dot(left_centered, left_centered))
                * float(np.dot(right_centered, right_centered))
            )
            correlation = math.nan
            if denominator > 0:
                correlation = float(
                    np.dot(left_centered, right_centered) / denominator
                )
                correlation = min(1.0, max(-1.0, correlation))

            # 序列无波动时设为最远距离，不让其被错误合并。
            distance = 2.0 if math.isnan(correlation) else 1.0 - correlation
            correlations[left_position, right_position] = correlation
            correlations[right_position, left_position] = correlation
            distances[left_position, right_position] = distance
            distances[right_position, left_position] = distance
    return correlations, distances


def complete_linkage_clusters(
    index_codes: Sequence[str],
    distance_matrix: np.ndarray,
) -> dict[str, int]:
    if not index_codes:
        return {}
    if len(index_codes) == 1:
        return {index_codes[0]: 1}

    distance_threshold = 1.0 - CORRELATION_THRESHOLD
    condensed_distance = squareform(distance_matrix, checks=False)
    linkage_matrix = linkage(condensed_distance, method="complete")
    raw_labels = fcluster(
        linkage_matrix,
        t=distance_threshold,
        criterion="distance",
    )

    # SciPy的原始编号不作为业务含义；按每类最小指数代码重排，保证输出稳定。
    codes_by_raw_label: dict[int, list[str]] = defaultdict(list)
    for index_code, raw_label in zip(index_codes, raw_labels):
        codes_by_raw_label[int(raw_label)].append(index_code)
    ordered_groups = sorted(
        codes_by_raw_label.values(),
        key=lambda codes: min(codes),
    )
    cluster_by_code: dict[str, int] = {}
    for cluster_number, codes in enumerate(ordered_groups, start=1):
        for index_code in codes:
            cluster_by_code[index_code] = cluster_number

    # complete linkage在阈值切割后，同类任意两指数距离都不应超过阈值。
    position_by_code = {
        index_code: position for position, index_code in enumerate(index_codes)
    }
    for codes in ordered_groups:
        positions = [position_by_code[code] for code in codes]
        if len(positions) < 2:
            continue
        maximum_distance = max(
            distance_matrix[left, right]
            for left in positions
            for right in positions
        )
        if maximum_distance > distance_threshold + 1e-12:
            raise RuntimeError(
                "complete linkage聚类结果未满足距离阈值："
                f"{maximum_distance:.6f} > {distance_threshold:.6f}"
            )
    return cluster_by_code


def select_cluster_representatives(
    month: MonthInput,
    cluster_by_code: Mapping[str, int],
    average_turnover: Mapping[tuple[str, str], float],
) -> list[dict[str, str]]:
    rows_by_cluster: dict[int, list[Mapping[str, str]]] = defaultdict(list)
    for row in month.etf_rows:
        index_code = clean_text(row.get("对标指数代码"))
        rows_by_cluster[cluster_by_code[index_code]].append(row)

    index_count_by_cluster: dict[int, int] = {
        cluster_number: len(
            {
                clean_text(row.get("对标指数代码"))
                for row in rows
            }
        )
        for cluster_number, rows in rows_by_cluster.items()
    }

    selected: list[dict[str, str]] = []
    for cluster_number in sorted(rows_by_cluster):
        rows = rows_by_cluster[cluster_number]
        winner = min(
            rows,
            key=lambda row: (
                -average_turnover[
                    (month.file_name, clean_text(row.get("代码")))
                ],
                clean_text(row.get("代码")),
            ),
        )
        winner_code = clean_text(winner.get("代码"))
        output_row = dict(winner)
        output_row.update(
            {
                "过去20个交易日平均成交额": format(
                    average_turnover[(month.file_name, winner_code)], ".15g"
                ),
                "聚类编号": str(cluster_number),
                "聚类指数数量": str(index_count_by_cluster[cluster_number]),
            }
        )
        selected.append(output_row)
    return selected


def build_cluster_detail_rows(
    month: MonthInput,
    cluster_by_code: Mapping[str, int],
    average_turnover: Mapping[tuple[str, str], float],
    selected_rows: Sequence[Mapping[str, str]],
    index_codes: Sequence[str],
    correlation_matrix: np.ndarray,
) -> list[dict[str, str]]:
    """保留代表筛选前的所有指数，输出每个聚类的成员信息。"""

    representative_by_cluster = {
        int(clean_text(row.get("聚类编号"))): row
        for row in selected_rows
    }
    rows_by_cluster: dict[int, list[Mapping[str, str]]] = defaultdict(list)
    for row in month.etf_rows:
        index_code = clean_text(row.get("对标指数代码"))
        rows_by_cluster[cluster_by_code[index_code]].append(row)

    position_by_code = {
        index_code: position for position, index_code in enumerate(index_codes)
    }
    sortable_details: list[tuple[int, int, float, str, dict[str, str]]] = []
    for cluster_number in sorted(rows_by_cluster):
        members = rows_by_cluster[cluster_number]
        representative = representative_by_cluster[cluster_number]
        representative_etf_code = clean_text(representative.get("代码"))
        representative_index_code = clean_text(
            representative.get("对标指数代码")
        )
        member_index_codes = sorted(
            {
                clean_text(row.get("对标指数代码"))
                for row in members
            }
        )

        cluster_correlations = [
            float(
                correlation_matrix[
                    position_by_code[left_code],
                    position_by_code[right_code],
                ]
            )
            for left_position, left_code in enumerate(member_index_codes)
            for right_code in member_index_codes[left_position + 1 :]
        ]
        minimum_cluster_correlation = (
            min(cluster_correlations) if cluster_correlations else None
        )

        for member in members:
            etf_code = clean_text(member.get("代码"))
            index_code = clean_text(member.get("对标指数代码"))
            is_representative = etf_code == representative_etf_code
            member_average_turnover = average_turnover[
                (month.file_name, etf_code)
            ]
            representative_correlation = float(
                correlation_matrix[
                    position_by_code[index_code],
                    position_by_code[representative_index_code],
                ]
            )
            detail = {
                "聚类编号": str(cluster_number),
                "是否最终代表": "是" if is_representative else "否",
                "ETF代码": etf_code,
                "ETF名称": clean_text(member.get("名称")),
                "类别大类": clean_text(member.get("类别大类")),
                "对标指数代码": index_code,
                "对标指数": clean_text(member.get("对标指数")),
                "过去20个交易日平均成交额": format(
                    member_average_turnover,
                    ".15g",
                ),
                "与代表指数相关性": format(
                    representative_correlation,
                    ".15g",
                ),
                "聚类内最低相关性": (
                    ""
                    if minimum_cluster_correlation is None
                    else format(minimum_cluster_correlation, ".15g")
                ),
                "相关性阈值": format(
                    CORRELATION_THRESHOLD,
                    ".15g",
                ),
                "距离阈值": format(
                    1.0 - CORRELATION_THRESHOLD,
                    ".15g",
                ),
            }
            sortable_details.append(
                (
                    cluster_number,
                    0 if is_representative else 1,
                    -member_average_turnover,
                    etf_code,
                    detail,
                )
            )

    sortable_details.sort(key=lambda item: item[:4])
    return [item[4] for item in sortable_details]


def excel_cell_value(field_name: str, value: object) -> object:
    if field_name in {"日期", "上市日期"}:
        try:
            return parse_date(value, field_name)
        except ValueError:
            return clean_text(value)
    if field_name in {
        "规模",
        "成交量",
        "成交额",
        "溢价率",
        "过去20个交易日平均成交额",
        "聚类编号",
        "聚类指数数量",
        "相关性阈值",
        "距离阈值",
        "与代表指数相关性",
        "聚类内最低相关性",
    }:
        return parse_finite_number(value)
    return clean_text(value)


def write_month_workbook(
    path: Path,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, str]],
    month: MonthInput,
    index_codes: Sequence[str],
    correlation_matrix: np.ndarray,
    cluster_detail_rows: Sequence[Mapping[str, str]],
) -> None:
    output_fieldnames = list(fieldnames) + ADDED_OUTPUT_COLUMNS
    index_name_by_code: dict[str, str] = {}
    for row in month.etf_rows:
        index_code = clean_text(row.get("对标指数代码"))
        index_name_by_code.setdefault(
            index_code,
            clean_text(row.get("对标指数")),
        )

    workbook = Workbook()
    selection_sheet = workbook.active
    selection_sheet.title = "动态指数池"
    correlation_sheet = workbook.create_sheet("相关性矩阵")
    cluster_detail_sheet = workbook.create_sheet("聚类明细")

    selection_sheet.append(output_fieldnames)
    for row in rows:
        selection_sheet.append(
            [
                excel_cell_value(field_name, row.get(field_name, ""))
                for field_name in output_fieldnames
            ]
        )

    dark_blue_fill = PatternFill("solid", fgColor="1F4E78")
    light_blue_fill = PatternFill("solid", fgColor="D9EAF7")
    white_bold_font = Font(color="FFFFFF", bold=True)
    dark_bold_font = Font(color="1F1F1F", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
    centered = Alignment(horizontal="center", vertical="center")

    selection_sheet.sheet_view.showGridLines = False
    selection_sheet.freeze_panes = "A2"
    selection_sheet.auto_filter.ref = selection_sheet.dimensions
    selection_sheet.row_dimensions[1].height = 28
    for cell in selection_sheet[1]:
        cell.fill = dark_blue_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row_cells in selection_sheet.iter_rows(
        min_row=2,
        max_row=selection_sheet.max_row,
    ):
        for cell in row_cells:
            cell.border = thin_border

    selection_widths = {
        "日期": 12,
        "代码": 14,
        "名称": 24,
        "上市日期": 12,
        "对标指数": 28,
        "benchmark": 34,
        "管理人": 28,
        "托管人": 28,
        "类别大类": 16,
        "类别细类": 18,
        "对标指数代码": 18,
        "过去20个交易日平均成交额": 24,
    }
    field_position = {
        field_name: position
        for position, field_name in enumerate(output_fieldnames, start=1)
    }
    for position, field_name in enumerate(output_fieldnames, start=1):
        selection_sheet.column_dimensions[get_column_letter(position)].width = (
            selection_widths.get(field_name, 16)
        )
    for field_name in ("日期", "上市日期"):
        position = field_position.get(field_name)
        if position is not None:
            for cell in selection_sheet.iter_cols(
                min_col=position,
                max_col=position,
                min_row=2,
                max_row=selection_sheet.max_row,
            ):
                for item in cell:
                    item.number_format = "yyyy-mm-dd"
    for field_name in (
        "规模",
        "成交量",
        "成交额",
        "过去20个交易日平均成交额",
    ):
        position = field_position.get(field_name)
        if position is not None:
            for cell in selection_sheet.iter_cols(
                min_col=position,
                max_col=position,
                min_row=2,
                max_row=selection_sheet.max_row,
            ):
                for item in cell:
                    item.number_format = "#,##0.00"
    for field_name in ("溢价率",):
        position = field_position.get(field_name)
        if position is not None:
            for cell in selection_sheet.iter_cols(
                min_col=position,
                max_col=position,
                min_row=2,
                max_row=selection_sheet.max_row,
            ):
                for item in cell:
                    item.number_format = "0.0000"
    if selection_sheet.max_row > 1:
        table_reference = (
            f"A1:{get_column_letter(selection_sheet.max_column)}"
            f"{selection_sheet.max_row}"
        )
        selection_table = Table(
            displayName="DynamicIndexPool",
            ref=table_reference,
        )
        selection_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        selection_sheet.add_table(selection_table)

    correlation_sheet.append(
        ["对标指数代码", "对标指数", *index_codes]
    )
    for row_position, index_code in enumerate(index_codes):
        correlation_sheet.append(
            [
                index_code,
                index_name_by_code.get(index_code, ""),
                *[
                    None if math.isnan(float(value)) else float(value)
                    for value in correlation_matrix[row_position]
                ],
            ]
        )

    correlation_sheet.sheet_view.showGridLines = False
    correlation_sheet.freeze_panes = "C2"
    correlation_sheet.row_dimensions[1].height = 42
    for cell in correlation_sheet[1]:
        cell.fill = dark_blue_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row_number in range(2, correlation_sheet.max_row + 1):
        for column_number in (1, 2):
            cell = correlation_sheet.cell(row_number, column_number)
            cell.fill = light_blue_fill
            cell.font = dark_bold_font
        for column_number in range(3, correlation_sheet.max_column + 1):
            cell = correlation_sheet.cell(row_number, column_number)
            cell.number_format = "0.0000"
            cell.alignment = centered

    correlation_sheet.column_dimensions["A"].width = 18
    correlation_sheet.column_dimensions["B"].width = 30
    for column_number in range(3, correlation_sheet.max_column + 1):
        correlation_sheet.column_dimensions[
            get_column_letter(column_number)
        ].width = 13
    if index_codes:
        matrix_range = (
            f"C2:{get_column_letter(correlation_sheet.max_column)}"
            f"{correlation_sheet.max_row}"
        )
        correlation_sheet.conditional_formatting.add(
            matrix_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0.8,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    cluster_detail_sheet.append(CLUSTER_DETAIL_COLUMNS)
    for detail in cluster_detail_rows:
        cluster_detail_sheet.append(
            [
                excel_cell_value(field_name, detail.get(field_name, ""))
                for field_name in CLUSTER_DETAIL_COLUMNS
            ]
        )

    cluster_detail_sheet.sheet_view.showGridLines = False
    cluster_detail_sheet.freeze_panes = "A2"
    cluster_detail_sheet.auto_filter.ref = cluster_detail_sheet.dimensions
    cluster_detail_sheet.row_dimensions[1].height = 32
    for cell in cluster_detail_sheet[1]:
        cell.fill = dark_blue_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    representative_flag_position = (
        CLUSTER_DETAIL_COLUMNS.index("是否最终代表") + 1
    )
    for row_number in range(2, cluster_detail_sheet.max_row + 1):
        for cell in cluster_detail_sheet[row_number]:
            cell.border = thin_border
        if (
            cluster_detail_sheet.cell(
                row_number,
                representative_flag_position,
            ).value
            == "是"
        ):
            representative_fill = PatternFill("solid", fgColor="C6E0B4")
            for representative_cell in cluster_detail_sheet[row_number]:
                representative_cell.fill = representative_fill
            representative_flag_cell = cluster_detail_sheet.cell(
                row_number,
                representative_flag_position,
            )
            representative_flag_cell.font = Font(color="006100", bold=True)
            representative_flag_cell.alignment = centered

    cluster_detail_widths = {
        "聚类编号": 12,
        "是否最终代表": 16,
        "ETF代码": 14,
        "ETF名称": 26,
        "类别大类": 16,
        "对标指数代码": 18,
        "对标指数": 28,
        "过去20个交易日平均成交额": 24,
        "与代表指数相关性": 20,
        "聚类内最低相关性": 20,
        "相关性阈值": 16,
        "距离阈值": 14,
    }
    cluster_detail_position = {
        field_name: position
        for position, field_name in enumerate(CLUSTER_DETAIL_COLUMNS, start=1)
    }
    for position, field_name in enumerate(CLUSTER_DETAIL_COLUMNS, start=1):
        cluster_detail_sheet.column_dimensions[
            get_column_letter(position)
        ].width = cluster_detail_widths[field_name]
    amount_position = cluster_detail_position["过去20个交易日平均成交额"]
    for cell in cluster_detail_sheet.iter_cols(
        min_col=amount_position,
        max_col=amount_position,
        min_row=2,
        max_row=cluster_detail_sheet.max_row,
    ):
        for item in cell:
            item.number_format = "#,##0.00"
    for field_name in (
        "与代表指数相关性",
        "聚类内最低相关性",
        "相关性阈值",
        "距离阈值",
    ):
        position = cluster_detail_position[field_name]
        for cell in cluster_detail_sheet.iter_cols(
            min_col=position,
            max_col=position,
            min_row=2,
            max_row=cluster_detail_sheet.max_row,
        ):
            for item in cell:
                item.number_format = "0.0000"
    if cluster_detail_sheet.max_row > 1:
        cluster_detail_table = Table(
            displayName="ClusterDetails",
            ref=(
                f"A1:{get_column_letter(cluster_detail_sheet.max_column)}"
                f"{cluster_detail_sheet.max_row}"
            ),
        )
        cluster_detail_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        cluster_detail_sheet.add_table(cluster_detail_table)

    temp_path = path.with_name(f".{path.stem}.tmp.xlsx")
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()


def validate_parameters() -> None:
    if not any(
        math.isclose(CORRELATION_THRESHOLD, allowed, abs_tol=1e-12)
        for allowed in ALLOWED_CORRELATION_THRESHOLDS
    ):
        raise ValueError(
            "CORRELATION_THRESHOLD只能设为0.7、0.8或0.9"
        )
    if RETURN_TRADING_DAYS <= 0:
        raise ValueError("RETURN_TRADING_DAYS必须大于0")
    if TURNOVER_LOOKBACK_DAYS <= 0:
        raise ValueError("TURNOVER_LOOKBACK_DAYS必须大于0")


def remove_stale_outputs(expected_file_names: set[str]) -> int:
    stale_files = [
        path
        for path in OUTPUT_DIR.iterdir()
        if path.is_file()
        and path.suffix.lower() in {".csv", ".xlsx"}
        and path.name not in expected_file_names
    ]
    for path in stale_files:
        path.unlink()
    return len(stale_files)


def main() -> None:
    validate_parameters()
    fieldnames, months = discover_month_inputs()
    print(
        f"发现 {len(months)} 个月度ETF初筛池；"
        f"相关性阈值 {CORRELATION_THRESHOLD:g}，"
        f"距离阈值 {1.0 - CORRELATION_THRESHOLD:g}，"
        "聚类方法 complete linkage",
        flush=True,
    )

    print("正在从etf_data.csv确定交易日日历...", flush=True)
    available_dates = inspect_etf_data_dates(ETF_DATA_FILE)
    turnover_windows = build_turnover_windows(months, available_dates)
    print("正在计算入选ETF过去20个交易日平均成交额...", flush=True)
    average_turnover = calculate_average_turnover(
        ETF_DATA_FILE,
        months,
        turnover_windows,
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for position, month in enumerate(months, start=1):
        returns_by_code = read_month_index_returns(month)
        index_codes = sorted(month.index_codes)
        correlation_matrix, distance_matrix = (
            calculate_correlation_and_distance_matrices(
                index_codes,
                returns_by_code,
            )
        )
        cluster_by_code = complete_linkage_clusters(
            index_codes,
            distance_matrix,
        )
        selected_rows = select_cluster_representatives(
            month,
            cluster_by_code,
            average_turnover,
        )
        cluster_detail_rows = build_cluster_detail_rows(
            month,
            cluster_by_code,
            average_turnover,
            selected_rows,
            index_codes,
            correlation_matrix,
        )
        output_path = OUTPUT_DIR / month.file_name
        write_month_workbook(
            output_path,
            fieldnames,
            selected_rows,
            month,
            index_codes,
            correlation_matrix,
            cluster_detail_rows,
        )
        print(
            f"{month.selection_date}：{len(index_codes)} 个指数 -> "
            f"{len(selected_rows)} 个聚类代表，"
            f"已保存动态指数池、完整相关性矩阵和聚类明细 "
            f"{output_path} ({position}/{len(months)})",
            flush=True,
        )

    removed_count = remove_stale_outputs(
        {month.file_name for month in months}
    )
    if removed_count:
        print(f"已清理 {removed_count} 个旧聚类输出文件", flush=True)

    print(
        f"完成：共生成 {len(months)} 个月底XLSX，"
        "第一张表为动态指数池，第二张表为完整相关性矩阵，"
        "第三张表为聚类明细，"
        f"输出目录：{OUTPUT_DIR}",
        flush=True,
    )


if __name__ == "__main__":
    main()

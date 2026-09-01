#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
对月底动态指数池中的指数逐日计算三种趋势得分。

计算方法：
1. 对指数过去N个有效收盘价取自然对数；
2. 用 t = 0, 1, ..., N-1 回归 log(P) = a + b * t；
3. 计算窗口收益率、窗口波动率和回归R²；
4. 同时计算三种得分：
   - 收益率×R²；
   - 收益率÷波动率；
   - 收益率÷波动率×R²；
5. 保留原“收益率×R²”的排名列，另外两种得分由回测脚本独立排名。

时间规则：
- 月末动态指数池使用当月月末已知数据生成；
- 为避免未来数据，该指数池从下一个ETF交易日开始生效；
- 在下一个月末指数池生效前，继续使用上一期指数池。

输入：
- outputs/etf_pool/clusters/threshold_<阈值>/reports/*.xlsx
  只读取第一张“动态指数池”；
- outputs/etf_trend_strategy/threshold_<阈值>/index_prices/*.csv
  合并对应聚类阈值下的代表指数连续收盘价；
- outputs/etf_data/etf_data.csv
  仅用于确定ETF市场交易日。

输出：
- outputs/etf_trend_strategy/threshold_<阈值>/factors/window_<N>/<年份>.csv

本脚本不请求任何数据接口，只使用已有outputs数据。
"""

from __future__ import annotations

import csv
import math
from bisect import bisect_right
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]

# ============================= 因子参数 =============================
# 聚类结果阈值，可改为0.7、0.8或0.9。
CLUSTER_CORRELATION_THRESHOLD = 0.9

# 趋势回归窗口，可改为20、40或60。
TREND_WINDOW = 20

START_YEAR = 2021
END_YEAR = 2026

# 跨境指数的交易日与ETF市场可能不一致。允许使用不晚于当日的
# 最新指数收盘价，但收盘价距当日超过该天数时因子记为数据不足。
MAX_PRICE_STALENESS_CALENDAR_DAYS = 7
# ====================================================================

ALLOWED_CLUSTER_THRESHOLDS = (0.7, 0.8, 0.9)
ALLOWED_TREND_WINDOWS = (20, 40, 60)

CLUSTER_POOL_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_pool"
    / "clusters"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "reports"
)
INDEX_RAW_DATA_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_trend_strategy"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "index_prices"
)
ETF_DATA_FILE = PROJECT_ROOT / "outputs" / "etf_data" / "etf_data.csv"
OUTPUT_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_trend_strategy"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "factors"
    / f"window_{TREND_WINDOW}"
)

POOL_SHEET_NAME = "动态指数池"
POOL_REQUIRED_COLUMNS = {
    "日期",
    "代码",
    "名称",
    "对标指数",
    "对标指数代码",
    "过去20个交易日平均成交额",
}
INDEX_RAW_REQUIRED_COLUMNS = {
    "收益日期",
    "对标指数代码",
    "对标指数",
    "收盘价",
}
ETF_DATA_REQUIRED_COLUMNS = {"日期"}

OUTPUT_COLUMNS = [
    "日期",
    "指数池日期",
    "ETF代码",
    "ETF名称",
    "对标指数代码",
    "对标指数",
    "趋势窗口",
    "窗口起始日",
    "窗口结束日",
    "窗口起始收盘价",
    "窗口结束收盘价",
    "log价格回归斜率",
    "窗口收益率",
    "窗口波动率",
    "R平方",
    "趋势质量因子",
    "风险调整趋势得分",
    "风险调整R平方趋势得分",
    "因子排名",
    "排名百分比",
]


@dataclass(frozen=True)
class PoolMember:
    etf_code: str
    etf_name: str
    index_code: str
    index_name: str


@dataclass(frozen=True)
class PoolSnapshot:
    selection_date: date
    members: tuple[PoolMember, ...]
    source_file: Path


@dataclass(frozen=True)
class TrendResult:
    window_start_date: date
    window_end_date: date
    window_start_close: float
    window_end_close: float
    window_return: float
    window_volatility: float
    slope: float
    r_squared: float
    trend_factor: float
    risk_adjusted_trend_factor: float
    risk_adjusted_r_squared_trend_factor: float


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "--", "None", "null", "NULL"} else text


def parse_date(value: object, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    for date_format in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name}不是有效日期：{text!r}")


def parse_positive_float(value: object, field_name: str) -> float:
    text = clean_text(value).replace(",", "")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"{field_name}不是有效数字：{text!r}") from exc
    if not math.isfinite(number) or number <= 0:
        raise ValueError(f"{field_name}必须是正数：{text!r}")
    return number


def validate_parameters() -> None:
    if not any(
        math.isclose(
            CLUSTER_CORRELATION_THRESHOLD,
            allowed,
            abs_tol=1e-12,
        )
        for allowed in ALLOWED_CLUSTER_THRESHOLDS
    ):
        raise ValueError(
            "CLUSTER_CORRELATION_THRESHOLD只能设为0.7、0.8或0.9"
        )
    if TREND_WINDOW not in ALLOWED_TREND_WINDOWS:
        raise ValueError("TREND_WINDOW只能设为20、40或60")
    if START_YEAR > END_YEAR:
        raise ValueError("START_YEAR不能晚于END_YEAR")
    if MAX_PRICE_STALENESS_CALENDAR_DAYS < 0:
        raise ValueError("MAX_PRICE_STALENESS_CALENDAR_DAYS不能小于0")


def discover_pool_snapshots() -> list[PoolSnapshot]:
    if not CLUSTER_POOL_DIR.exists():
        raise FileNotFoundError(f"找不到月底动态指数池：{CLUSTER_POOL_DIR}")

    files = sorted(CLUSTER_POOL_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"动态指数池目录没有XLSX：{CLUSTER_POOL_DIR}")

    snapshots: list[PoolSnapshot] = []
    for path in files:
        try:
            file_date = datetime.strptime(path.stem, "%Y_%m_%d").date()
        except ValueError as exc:
            raise ValueError(f"动态指数池文件名日期无效：{path.name}") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if POOL_SHEET_NAME not in workbook.sheetnames:
                raise ValueError(f"{path.name}缺少工作表：{POOL_SHEET_NAME}")
            sheet = workbook[POOL_SHEET_NAME]
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise ValueError(f"{path.name}的{POOL_SHEET_NAME}为空") from exc
            headers = [clean_text(value) for value in header_row]
            missing_columns = POOL_REQUIRED_COLUMNS - set(headers)
            if missing_columns:
                raise ValueError(
                    f"{path.name}的{POOL_SHEET_NAME}缺少列："
                    f"{sorted(missing_columns)}"
                )
            position = {header: index for index, header in enumerate(headers)}

            members_by_index: dict[str, PoolMember] = {}
            for row_number, values in enumerate(rows, start=2):
                if not any(value is not None for value in values):
                    continue
                row_date = parse_date(values[position["日期"]], "日期")
                if row_date != file_date:
                    raise ValueError(
                        f"{path.name}第{row_number}行日期{row_date}"
                        "与文件名不一致"
                    )
                member = PoolMember(
                    etf_code=clean_text(values[position["代码"]]),
                    etf_name=clean_text(values[position["名称"]]),
                    index_code=clean_text(values[position["对标指数代码"]]),
                    index_name=clean_text(values[position["对标指数"]]),
                )
                if not member.etf_code or not member.index_code:
                    raise ValueError(f"{path.name}第{row_number}行ETF或指数代码为空")
                if member.index_code in members_by_index:
                    raise ValueError(
                        f"{path.name}存在重复代表指数：{member.index_code}"
                    )
                members_by_index[member.index_code] = member
        finally:
            workbook.close()

        if not members_by_index:
            raise ValueError(f"{path.name}的{POOL_SHEET_NAME}没有指数")
        snapshots.append(
            PoolSnapshot(
                selection_date=file_date,
                members=tuple(
                    members_by_index[index_code]
                    for index_code in sorted(members_by_index)
                ),
                source_file=path,
            )
        )

    snapshot_dates = [snapshot.selection_date for snapshot in snapshots]
    if len(snapshot_dates) != len(set(snapshot_dates)):
        raise ValueError("动态指数池存在重复月份")
    return snapshots


def read_etf_trading_calendar() -> list[date]:
    if not ETF_DATA_FILE.exists():
        raise FileNotFoundError(f"找不到ETF数据：{ETF_DATA_FILE}")

    trading_dates: set[date] = set()
    with ETF_DATA_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing_columns = ETF_DATA_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing_columns:
            raise ValueError(f"etf_data.csv缺少列：{sorted(missing_columns)}")
        for row in reader:
            current_date = parse_date(row.get("日期"), "日期")
            if START_YEAR <= current_date.year <= END_YEAR:
                trading_dates.add(current_date)

    if not trading_dates:
        raise ValueError(f"{START_YEAR}至{END_YEAR}年没有ETF交易日")
    return sorted(trading_dates)


def load_index_close_history(
    required_index_codes: set[str],
) -> tuple[dict[str, list[date]], dict[str, list[float]], dict[str, str]]:
    if not INDEX_RAW_DATA_DIR.exists():
        raise FileNotFoundError(f"找不到指数历史数据：{INDEX_RAW_DATA_DIR}")

    files = sorted(INDEX_RAW_DATA_DIR.glob("*.csv"))
    if not files:
        raise FileNotFoundError(f"指数历史目录没有CSV：{INDEX_RAW_DATA_DIR}")

    closes_by_code: dict[str, dict[date, float]] = defaultdict(dict)
    index_names: dict[str, str] = {}
    for path in files:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            missing_columns = INDEX_RAW_REQUIRED_COLUMNS - set(
                reader.fieldnames or []
            )
            if missing_columns:
                raise ValueError(f"{path.name}缺少列：{sorted(missing_columns)}")
            for row_number, row in enumerate(reader, start=2):
                index_code = clean_text(row.get("对标指数代码"))
                if index_code not in required_index_codes:
                    continue
                price_date = parse_date(row.get("收益日期"), "收益日期")
                close = parse_positive_float(row.get("收盘价"), "收盘价")
                existing = closes_by_code[index_code].get(price_date)
                if existing is not None and not math.isclose(
                    existing,
                    close,
                    rel_tol=1e-10,
                    abs_tol=1e-10,
                ):
                    raise ValueError(
                        f"{index_code} {price_date}在不同月度文件的"
                        f"收盘价不一致：{existing} vs {close}"
                    )
                closes_by_code[index_code][price_date] = close
                name = clean_text(row.get("对标指数"))
                if name:
                    index_names.setdefault(index_code, name)

    dates_by_code: dict[str, list[date]] = {}
    prices_by_code: dict[str, list[float]] = {}
    for index_code in sorted(required_index_codes):
        dated_prices = closes_by_code.get(index_code, {})
        ordered_dates = sorted(dated_prices)
        dates_by_code[index_code] = ordered_dates
        prices_by_code[index_code] = [
            dated_prices[current_date] for current_date in ordered_dates
        ]
    return dates_by_code, prices_by_code, index_names


def calculate_trend(
    current_date: date,
    price_dates: Sequence[date],
    prices: Sequence[float],
) -> tuple[TrendResult | None, str]:
    end_position = bisect_right(price_dates, current_date)
    if end_position < TREND_WINDOW:
        return None, f"截至当日不足{TREND_WINDOW}个有效收盘价"

    window_dates = price_dates[end_position - TREND_WINDOW : end_position]
    window_prices = np.asarray(
        prices[end_position - TREND_WINDOW : end_position],
        dtype=float,
    )
    last_price_date = window_dates[-1]
    stale_days = (current_date - last_price_date).days
    if stale_days > MAX_PRICE_STALENESS_CALENDAR_DAYS:
        return None, f"最新收盘价滞后{stale_days}个自然日"

    log_prices = np.log(window_prices)
    time_index = np.arange(TREND_WINDOW, dtype=float)
    design_matrix = np.column_stack(
        (np.ones(TREND_WINDOW, dtype=float), time_index)
    )
    intercept, slope = np.linalg.lstsq(
        design_matrix,
        log_prices,
        rcond=None,
    )[0]
    fitted = intercept + slope * time_index
    residual_sum_squares = float(np.sum((log_prices - fitted) ** 2))
    total_sum_squares = float(
        np.sum((log_prices - float(np.mean(log_prices))) ** 2)
    )
    if total_sum_squares <= np.finfo(float).eps:
        r_squared = 1.0
    else:
        r_squared = 1.0 - residual_sum_squares / total_sum_squares
        r_squared = min(1.0, max(0.0, r_squared))

    window_return = float(window_prices[-1] / window_prices[0] - 1.0)
    daily_returns = window_prices[1:] / window_prices[:-1] - 1.0
    daily_volatility = float(np.std(daily_returns, ddof=1))
    window_volatility = daily_volatility * math.sqrt(len(daily_returns))
    if (
        not math.isfinite(window_volatility)
        or window_volatility <= np.finfo(float).eps
    ):
        risk_adjusted_trend_factor = 0.0
    else:
        risk_adjusted_trend_factor = window_return / window_volatility

    return (
        TrendResult(
            window_start_date=window_dates[0],
            window_end_date=window_dates[-1],
            window_start_close=float(window_prices[0]),
            window_end_close=float(window_prices[-1]),
            window_return=window_return,
            window_volatility=window_volatility,
            slope=float(slope),
            r_squared=r_squared,
            trend_factor=window_return * r_squared,
            risk_adjusted_trend_factor=risk_adjusted_trend_factor,
            risk_adjusted_r_squared_trend_factor=(
                risk_adjusted_trend_factor * r_squared
            ),
        ),
        "有效",
    )


def active_pool_snapshot(
    current_date: date,
    snapshots: Sequence[PoolSnapshot],
    snapshot_dates: Sequence[date],
) -> PoolSnapshot | None:
    # 严格小于当日：月末结果从下一交易日起生效。
    position = bisect_right(snapshot_dates, current_date) - 1
    while position >= 0 and snapshot_dates[position] >= current_date:
        position -= 1
    return None if position < 0 else snapshots[position]


def build_daily_rows(
    current_date: date,
    snapshot: PoolSnapshot,
    dates_by_code: Mapping[str, Sequence[date]],
    prices_by_code: Mapping[str, Sequence[float]],
) -> list[dict[str, object]]:
    calculated: list[tuple[PoolMember, TrendResult]] = []
    invalid: list[tuple[PoolMember, str]] = []
    for member in snapshot.members:
        result, status = calculate_trend(
            current_date,
            dates_by_code.get(member.index_code, ()),
            prices_by_code.get(member.index_code, ()),
        )
        if result is None:
            invalid.append((member, status))
        else:
            calculated.append((member, result))

    calculated.sort(
        key=lambda item: (-item[1].trend_factor, item[0].index_code)
    )
    valid_count = len(calculated)
    rows: list[dict[str, object]] = []
    for rank, (member, result) in enumerate(calculated, start=1):
        rows.append(
            {
                "日期": current_date.isoformat(),
                "指数池日期": snapshot.selection_date.isoformat(),
                "ETF代码": member.etf_code,
                "ETF名称": member.etf_name,
                "对标指数代码": member.index_code,
                "对标指数": member.index_name,
                "趋势窗口": TREND_WINDOW,
                "窗口起始日": result.window_start_date.isoformat(),
                "窗口结束日": result.window_end_date.isoformat(),
                "窗口起始收盘价": format(result.window_start_close, ".15g"),
                "窗口结束收盘价": format(result.window_end_close, ".15g"),
                "窗口收益率": format(result.window_return, ".15g"),
                "窗口波动率": format(result.window_volatility, ".15g"),
                "log价格回归斜率": format(result.slope, ".15g"),
                "R平方": format(result.r_squared, ".15g"),
                "趋势质量因子": format(result.trend_factor, ".15g"),
                "风险调整趋势得分": format(
                    result.risk_adjusted_trend_factor,
                    ".15g",
                ),
                "风险调整R平方趋势得分": format(
                    result.risk_adjusted_r_squared_trend_factor,
                    ".15g",
                ),
                "因子排名": rank,
                "排名百分比": format(rank / valid_count, ".15g"),
                "当日有效指数数量": valid_count,
                "数据状态": "有效",
            }
        )

    for member, status in sorted(invalid, key=lambda item: item[0].index_code):
        rows.append(
            {
                "日期": current_date.isoformat(),
                "指数池日期": snapshot.selection_date.isoformat(),
                "ETF代码": member.etf_code,
                "ETF名称": member.etf_name,
                "对标指数代码": member.index_code,
                "对标指数": member.index_name,
                "趋势窗口": TREND_WINDOW,
                "数据状态": status,
            }
        )
    return rows


def write_year_output(year: int, rows: Sequence[Mapping[str, object]]) -> Path:
    output_path = OUTPUT_DIR / f"{year}.csv"
    temp_path = output_path.with_name(f".{output_path.name}.tmp")
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=OUTPUT_COLUMNS,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(rows)
        temp_path.replace(output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return output_path


def main() -> None:
    validate_parameters()
    snapshots = discover_pool_snapshots()
    snapshot_dates = [snapshot.selection_date for snapshot in snapshots]
    required_index_codes = {
        member.index_code
        for snapshot in snapshots
        for member in snapshot.members
    }

    print(
        f"聚类阈值 {CLUSTER_CORRELATION_THRESHOLD:g}，"
        f"趋势窗口 {TREND_WINDOW} 个交易日，"
        f"共 {len(snapshots)} 期动态指数池。",
        flush=True,
    )
    latest_completed_pool_date = min(
        snapshot_dates[-1],
        date(END_YEAR, 12, 31),
    )
    trading_dates = [
        current_date
        for current_date in read_etf_trading_calendar()
        if current_date <= latest_completed_pool_date
    ]
    if not trading_dates:
        raise ValueError(
            f"截至最新已完成指数池日期{latest_completed_pool_date}没有ETF交易日"
        )
    print(
        f"计算截止日：{latest_completed_pool_date}"
        "（不计算尚未完成月末筛选的月份）。",
        flush=True,
    )
    dates_by_code, prices_by_code, _ = load_index_close_history(
        required_index_codes
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows_by_year: dict[int, list[dict[str, object]]] = defaultdict(list)
    valid_row_count = 0
    invalid_row_count = 0
    for position, current_date in enumerate(trading_dates, start=1):
        snapshot = active_pool_snapshot(
            current_date,
            snapshots,
            snapshot_dates,
        )
        if snapshot is None:
            continue
        daily_rows = build_daily_rows(
            current_date,
            snapshot,
            dates_by_code,
            prices_by_code,
        )
        rows_by_year[current_date.year].extend(daily_rows)
        valid_row_count += sum(
            row["数据状态"] == "有效" for row in daily_rows
        )
        invalid_row_count += sum(
            row["数据状态"] != "有效" for row in daily_rows
        )
        print(
            f"正在计算 {current_date} 趋势因子... "
            f"({position}/{len(trading_dates)})",
            flush=True,
        )

    expected_files: set[str] = set()
    for year in range(START_YEAR, END_YEAR + 1):
        year_rows = rows_by_year.get(year, [])
        output_path = write_year_output(year, year_rows)
        expected_files.add(output_path.name)
        print(f"✅ 已保存 {year} 年因子数据：{output_path}", flush=True)

    for old_file in OUTPUT_DIR.glob("*.csv"):
        if old_file.name not in expected_files:
            old_file.unlink()

    print(
        f"完成：有效因子 {valid_row_count} 条，"
        f"数据不足 {invalid_row_count} 条，"
        f"输出目录：{OUTPUT_DIR}",
        flush=True,
    )


if __name__ == "__main__":
    main()

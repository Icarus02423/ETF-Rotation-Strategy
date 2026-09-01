#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETF趋势轮动策略回测。

策略：
1. 分别按两种趋势得分降序选择前10%的代表指数；
2. 排名后仅保留当前趋势窗口收益率大于0的指数，不向后补选；
3. 对每个保留指数，在当日所有跟踪该指数且成交量大于0的ETF中，
   选择成交量最大者；成交量相同时依次比较成交额、规模和ETF代码；
4. 单个指数权重按过滤前计划入选数量等权，空缺权重保留为现金；
5. 分别按信号日收盘价和下一交易日VWAP成交；
6. 支持两种调仓模式：固定x日全组合调仓，或将资金分成x个独立账户，
   每天轮换一个账户、每个账户持有x个交易日；
7. 买入和卖出均收取0.1%的单边交易成本。

输入：
- outputs/etf_trend_strategy/threshold_<阈值>/factors/window_<窗口>/YYYY.csv
- outputs/etf_data/etf_data.csv

输出：
- 固定x日：post_rank_positive_return/rebalance_<x>d/<成交方式>/
- x账户错峰：post_rank_positive_return/staggered_<x>_accounts_hold_<x>d/<成交方式>/
  每个交易模式独立输出年度指标、总回测指标、合并持仓、时序和账户明细五个Excel，
  以及累计净值、换手率、累计交易成本和策略容量四张图。
"""

from __future__ import annotations

import csv
import math
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from matplotlib.ticker import PercentFormatter
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]

# ============================= 回测参数 =============================
CLUSTER_CORRELATION_THRESHOLD = 0.7
TREND_WINDOW = 60
# 默认一次运行两种得分。若以后只想跑其中一种，可只保留对应英文键。
SCORE_METHODS_TO_RUN = ("return_r2", "return_vol")
TOP_PERCENT = 0.10
# "rebalance"：全组合每x个交易日调仓；"staggered"：x个账户逐日错峰持有x日。
REBALANCE_MODE = "staggered"
ACCOUNT_REBALANCE_INTERVAL = 12
ACCOUNT_COUNT = (
    1 if REBALANCE_MODE == "rebalance" else ACCOUNT_REBALANCE_INTERVAL
)
MIN_WINDOW_RETURN = 0.0
STRATEGY_VARIANT_DIR = "post_rank_positive_return"
ACCOUNT_VARIANT_DIR = (
    f"rebalance_{ACCOUNT_REBALANCE_INTERVAL}d"
    if REBALANCE_MODE == "rebalance"
    else (
        f"staggered_{ACCOUNT_REBALANCE_INTERVAL}d"
    )
)
TRANSACTION_COST_RATE = 0.001
ANNUAL_TRADING_DAYS = 252
ANNUAL_RISK_FREE_RATE = 0.015
INITIAL_NAV = 1.0
CAPACITY_DAILY_AMOUNT_RATIO = 0.10
CAPACITY_DESCENDING_QUANTILE = 0.95
# ====================================================================

ALLOWED_CLUSTER_THRESHOLDS = (0.7, 0.8, 0.9)
ALLOWED_TREND_WINDOWS = (20, 40, 60)
SCORE_COLUMNS = {
    "return_r2": "趋势质量因子",
    "return_vol": "风险调整趋势得分",
}
SCORE_LABELS = {
    "return_r2": "收益率×R平方",
    "return_vol": "收益率÷波动率",
}
FACTOR_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_trend_strategy"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "factors"
    / f"window_{TREND_WINDOW}"
)
ETF_DATA_FILE = PROJECT_ROOT / "outputs" / "etf_data" / "etf_data.csv"
BACKTEST_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_trend_strategy"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "backtests"
    / f"window_{TREND_WINDOW}"
)

FACTOR_REQUIRED_COLUMNS = {
    "日期",
    "对标指数代码",
    "对标指数",
    "窗口收益率",
} | set(SCORE_COLUMNS.values())
ETF_REQUIRED_COLUMNS = {
    "日期",
    "代码",
    "名称",
    "上市日期",
    "对标指数代码",
    "对标指数",
    "规模",
    "成交量",
    "成交额",
    "收盘价",
    "VWAP",
}


@dataclass(frozen=True)
class FactorMember:
    signal_date: date
    index_code: str
    index_name: str
    trend_factor: float
    window_return: float
    factor_rank: int = 0


@dataclass(frozen=True)
class DailyFactorSelection:
    signal_date: date
    planned_index_count: int
    members: tuple[FactorMember, ...]

    @property
    def filtered_index_count(self) -> int:
        return self.planned_index_count - len(self.members)


@dataclass(frozen=True)
class CandidateEtf:
    code: str
    name: str
    index_code: str
    index_name: str
    volume: float
    amount: float
    scale: float


@dataclass(frozen=True)
class TargetMember:
    signal_date: date
    index_code: str
    index_name: str
    etf_code: str
    etf_name: str
    target_weight: float
    trend_factor: float
    factor_rank: int
    selection_volume: float
    selection_amount: float
    selection_scale: float


@dataclass(frozen=True)
class DailyTarget:
    signal_date: date
    planned_index_count: int
    selected_index_count: int
    members: tuple[TargetMember, ...]

    @property
    def filtered_index_count(self) -> int:
        return self.planned_index_count - self.selected_index_count

    @property
    def unmapped_index_count(self) -> int:
        return self.selected_index_count - len(self.members)


@dataclass(frozen=True)
class PricePoint:
    close: float | None
    vwap: float | None
    amount: float | None


@dataclass
class Position:
    shares: float
    index_code: str
    index_name: str
    etf_name: str
    signal_date: date
    trend_factor: float
    factor_rank: int
    selection_volume: float
    target_weight: float


@dataclass
class AccountState:
    account_id: int
    positions: dict[str, Position]
    cash: float
    has_been_built: bool = False
    last_signal_date: date | None = None
    last_rebalance_date: date | None = None
    next_rebalance_date: date | None = None


@dataclass(frozen=True)
class HoldingRecord:
    current_date: date
    signal_date: date
    index_code: str
    index_name: str
    trend_factor: float
    factor_rank: int
    etf_code: str
    etf_name: str
    selection_volume: float
    target_weight: float
    actual_weight: float


@dataclass(frozen=True)
class AccountDailyRecord:
    current_date: date
    account_id: int
    account_nav: float
    etf_market_value: float
    cash: float
    cash_weight: float
    rebalance_attempted: bool
    rebalance_succeeded: bool
    signal_date: date | None
    last_rebalance_date: date | None
    next_rebalance_date: date | None


@dataclass(frozen=True)
class AccountHoldingRecord:
    current_date: date
    account_id: int
    signal_date: date
    etf_code: str
    etf_name: str
    market_value: float
    account_weight: float
    total_portfolio_weight: float


@dataclass(frozen=True)
class AccountTradeRecord:
    execution_date: date
    account_id: int
    signal_date: date | None
    etf_code: str
    etf_name: str
    direction: str
    before_value: float
    target_value: float
    trade_amount: float
    transaction_cost: float


@dataclass(frozen=True)
class NavRecord:
    current_date: date
    signal_date: date | None
    pre_trade_nav: float
    nav: float
    gross_return: float
    daily_return: float
    drawdown: float
    buy_ratio: float
    sell_ratio: float
    bilateral_ratio: float
    one_way_turnover: float
    transaction_cost: float
    transaction_cost_rate: float
    cumulative_cost_rate: float
    capacity: float | None
    holding_count: int
    cash_weight: float
    initial_build: bool
    rebalance_account_id: int | None
    rebalance_attempted: bool
    rebalance_succeeded: bool
    skip_reason: str
    selected_index_count: int
    unmapped_index_count: int
    missing_price_index_count: int


@dataclass(frozen=True)
class TradeDetail:
    etf_code: str
    etf_name: str
    direction: str
    before_value: float
    target_value: float
    trade_amount: float
    transaction_cost: float


@dataclass(frozen=True)
class RebalanceResult:
    positions: dict[str, Position]
    cash: float
    pre_trade_nav: float
    buy_amount: float
    sell_amount: float
    transaction_cost: float
    trade_details: tuple[TradeDetail, ...]
    succeeded: bool
    skip_reason: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "--", "None", "null", "NULL"} else text


def parse_date(value: object, field_name: str) -> date:
    text = clean_text(value)[:10]
    for date_format in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name}不是有效日期：{value!r}")


def finite_float(value: object) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def positive_float(value: object) -> float | None:
    number = finite_float(value)
    return number if number is not None and number > 0 else None


def validate_parameters() -> None:
    if not any(
        math.isclose(CLUSTER_CORRELATION_THRESHOLD, allowed, abs_tol=1e-12)
        for allowed in ALLOWED_CLUSTER_THRESHOLDS
    ):
        raise ValueError("CLUSTER_CORRELATION_THRESHOLD只能设为0.7、0.8或0.9")
    if TREND_WINDOW not in ALLOWED_TREND_WINDOWS:
        raise ValueError("TREND_WINDOW只能设为20、40或60")
    invalid_score_methods = [
        method for method in SCORE_METHODS_TO_RUN if method not in SCORE_COLUMNS
    ]
    if not SCORE_METHODS_TO_RUN or invalid_score_methods:
        raise ValueError(
            "SCORE_METHODS_TO_RUN包含无效得分方法："
            f"{invalid_score_methods}"
        )
    if len(set(SCORE_METHODS_TO_RUN)) != len(SCORE_METHODS_TO_RUN):
        raise ValueError("SCORE_METHODS_TO_RUN不能包含重复得分方法")
    if not 0 < TOP_PERCENT <= 1:
        raise ValueError("TOP_PERCENT必须在0到1之间")
    if REBALANCE_MODE not in {"rebalance", "staggered"}:
        raise ValueError("REBALANCE_MODE只能设为rebalance或staggered")
    if not isinstance(ACCOUNT_COUNT, int) or ACCOUNT_COUNT <= 0:
        raise ValueError("ACCOUNT_COUNT必须是正整数")
    if (
        not isinstance(ACCOUNT_REBALANCE_INTERVAL, int)
        or ACCOUNT_REBALANCE_INTERVAL <= 0
    ):
        raise ValueError("ACCOUNT_REBALANCE_INTERVAL必须是正整数")
    expected_account_count = (
        1 if REBALANCE_MODE == "rebalance" else ACCOUNT_REBALANCE_INTERVAL
    )
    if ACCOUNT_COUNT != expected_account_count:
        raise ValueError(
            "ACCOUNT_COUNT与REBALANCE_MODE、ACCOUNT_REBALANCE_INTERVAL不一致"
        )
    if not math.isfinite(MIN_WINDOW_RETURN):
        raise ValueError("MIN_WINDOW_RETURN必须是有限数值")
    if TRANSACTION_COST_RATE < 0:
        raise ValueError("TRANSACTION_COST_RATE不能为负数")
    if not 0 < CAPACITY_DAILY_AMOUNT_RATIO <= 1:
        raise ValueError("CAPACITY_DAILY_AMOUNT_RATIO必须在0到1之间")
    if not 0 <= CAPACITY_DESCENDING_QUANTILE <= 1:
        raise ValueError("CAPACITY_DESCENDING_QUANTILE必须在0到1之间")
    if ANNUAL_TRADING_DAYS <= 0 or INITIAL_NAV <= 0:
        raise ValueError("年化交易日和初始净值必须大于0")


def read_daily_top_factors(
    score_column: str,
) -> dict[date, DailyFactorSelection]:
    if not FACTOR_DIR.exists():
        raise FileNotFoundError(f"找不到趋势因子目录：{FACTOR_DIR}")
    files = sorted(
        path
        for path in FACTOR_DIR.glob("*.csv")
        if not path.name.startswith(".") and path.stem.isdigit()
    )
    if not files:
        raise FileNotFoundError(f"趋势因子目录没有年度CSV：{FACTOR_DIR}")

    members_by_date: dict[date, dict[str, FactorMember]] = defaultdict(dict)
    for path in files:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            missing = FACTOR_REQUIRED_COLUMNS - set(reader.fieldnames or [])
            if missing:
                raise ValueError(f"{path.name}缺少列：{sorted(missing)}")
            for row_number, row in enumerate(reader, start=2):
                factor = finite_float(row.get(score_column))
                window_return = finite_float(row.get("窗口收益率"))
                index_code = clean_text(row.get("对标指数代码"))
                if factor is None or window_return is None or not index_code:
                    continue
                signal_date = parse_date(row.get("日期"), "日期")
                if index_code in members_by_date[signal_date]:
                    raise ValueError(
                        f"{path.name}第{row_number}行出现重复指数："
                        f"{signal_date} {index_code}"
                    )
                members_by_date[signal_date][index_code] = FactorMember(
                    signal_date=signal_date,
                    index_code=index_code,
                    index_name=clean_text(row.get("对标指数")),
                    trend_factor=factor,
                    window_return=window_return,
                )

    daily_selections: dict[date, DailyFactorSelection] = {}
    for signal_date in sorted(members_by_date):
        valid_members = sorted(
            members_by_date[signal_date].values(),
            key=lambda member: (-member.trend_factor, member.index_code),
        )
        if not valid_members:
            continue
        planned_count = max(1, math.ceil(len(valid_members) * TOP_PERCENT))
        ranked_members = tuple(
            FactorMember(
                signal_date=member.signal_date,
                index_code=member.index_code,
                index_name=member.index_name,
                trend_factor=member.trend_factor,
                window_return=member.window_return,
                factor_rank=factor_rank,
            )
            for factor_rank, member in enumerate(valid_members, start=1)
        )
        top_members = ranked_members[:planned_count]
        filtered_members = tuple(
            member
            for member in top_members
            if member.window_return > MIN_WINDOW_RETURN
        )
        daily_selections[signal_date] = DailyFactorSelection(
            signal_date=signal_date,
            planned_index_count=planned_count,
            members=filtered_members,
        )
    if not daily_selections:
        raise ValueError("趋势因子文件中没有有效因子")
    return daily_selections


def candidate_sort_key(candidate: CandidateEtf) -> tuple[float, float, float, str]:
    return (-candidate.volume, -candidate.amount, -candidate.scale, candidate.code)


def build_daily_targets(
    daily_selections: Mapping[date, DailyFactorSelection],
    etf_data_file: Path = ETF_DATA_FILE,
) -> tuple[dict[date, DailyTarget], list[date]]:
    """扫描一次ETF总表，同时建立交易日历和信号日的ETF目标。"""

    if not etf_data_file.exists():
        raise FileNotFoundError(f"找不到ETF数据：{etf_data_file}")
    required_indexes = {
        signal_date: {member.index_code for member in selection.members}
        for signal_date, selection in daily_selections.items()
    }
    best_by_date_index: dict[tuple[date, str], CandidateEtf] = {}
    trading_calendar: set[date] = set()

    with etf_data_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = ETF_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing:
            raise ValueError(
                f"etf_data.csv缺少列：{sorted(missing)}；"
                "请先完成ETF收盘价和VWAP下载。"
            )
        for row in reader:
            row_date_text = clean_text(row.get("日期"))
            try:
                current_date = date.fromisoformat(row_date_text[:10])
            except ValueError:
                continue
            trading_calendar.add(current_date)
            index_code = clean_text(row.get("对标指数代码"))
            if index_code not in required_indexes.get(current_date, set()):
                continue
            code = clean_text(row.get("代码")).upper()
            volume = positive_float(row.get("成交量"))
            if not code or volume is None:
                continue
            listed_text = clean_text(row.get("上市日期"))
            if listed_text:
                try:
                    if parse_date(listed_text, "上市日期") > current_date:
                        continue
                except ValueError:
                    continue
            candidate = CandidateEtf(
                code=code,
                name=clean_text(row.get("名称")),
                index_code=index_code,
                index_name=clean_text(row.get("对标指数")),
                volume=volume,
                amount=finite_float(row.get("成交额")) or 0.0,
                scale=finite_float(row.get("规模")) or 0.0,
            )
            key = (current_date, index_code)
            existing = best_by_date_index.get(key)
            if existing is None or candidate_sort_key(candidate) < candidate_sort_key(existing):
                best_by_date_index[key] = candidate

    targets: dict[date, DailyTarget] = {}
    for signal_date, selection in sorted(daily_selections.items()):
        target_weight = 1.0 / selection.planned_index_count
        mapped: list[TargetMember] = []
        for member in selection.members:
            candidate = best_by_date_index.get((signal_date, member.index_code))
            if candidate is None:
                continue
            mapped.append(
                TargetMember(
                    signal_date=signal_date,
                    index_code=member.index_code,
                    index_name=member.index_name or candidate.index_name,
                    etf_code=candidate.code,
                    etf_name=candidate.name,
                    target_weight=target_weight,
                    trend_factor=member.trend_factor,
                    factor_rank=member.factor_rank,
                    selection_volume=candidate.volume,
                    selection_amount=candidate.amount,
                    selection_scale=candidate.scale,
                )
            )
        targets[signal_date] = DailyTarget(
            signal_date=signal_date,
            planned_index_count=selection.planned_index_count,
            selected_index_count=len(selection.members),
            members=tuple(mapped),
        )
    if not trading_calendar:
        raise ValueError("etf_data.csv中没有有效交易日期")
    return targets, sorted(trading_calendar)


def load_selected_prices(
    selected_codes: set[str],
    required_dates: set[date],
    etf_data_file: Path = ETF_DATA_FILE,
) -> dict[date, dict[str, PricePoint]]:
    """第二次扫描总表，只把真正可能交易的ETF价格载入内存。"""

    prices: dict[date, dict[str, PricePoint]] = defaultdict(dict)
    with etf_data_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = ETF_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"etf_data.csv缺少列：{sorted(missing)}")
        for row in reader:
            code = clean_text(row.get("代码")).upper()
            if code not in selected_codes:
                continue
            row_date_text = clean_text(row.get("日期"))[:10]
            try:
                current_date = date.fromisoformat(row_date_text)
            except ValueError:
                continue
            if current_date not in required_dates:
                continue
            if code in prices[current_date]:
                raise ValueError(f"etf_data.csv存在重复ETF日期：{current_date} {code}")
            prices[current_date][code] = PricePoint(
                close=positive_float(row.get("收盘价")),
                vwap=positive_float(row.get("VWAP")),
                amount=positive_float(row.get("成交额")),
            )
    return dict(prices)


def linear_quantile(values: Sequence[float], quantile: float) -> float | None:
    """按NumPy默认的线性插值口径计算分位数。"""

    valid_values = sorted(value for value in values if math.isfinite(value))
    if not valid_values:
        return None
    if len(valid_values) == 1:
        return valid_values[0]
    position = (len(valid_values) - 1) * quantile
    lower_index = math.floor(position)
    upper_index = math.ceil(position)
    if lower_index == upper_index:
        return valid_values[lower_index]
    fraction = position - lower_index
    return (
        valid_values[lower_index] * (1.0 - fraction)
        + valid_values[upper_index] * fraction
    )


def aggregate_target_weights(
    target: DailyTarget,
    execution_prices: Mapping[str, float],
    close_prices: Mapping[str, float],
) -> tuple[dict[str, float], dict[str, TargetMember]]:
    """仅保留当日同时有成交价和收盘估值价的目标ETF，其余权重留作现金。"""

    weights: dict[str, float] = defaultdict(float)
    metadata: dict[str, TargetMember] = {}
    for member in target.members:
        if member.etf_code not in execution_prices or member.etf_code not in close_prices:
            continue
        weights[member.etf_code] += member.target_weight
        metadata.setdefault(member.etf_code, member)
    return dict(weights), metadata


def rebalance_portfolio(
    positions: dict[str, Position],
    cash: float,
    target_weights: Mapping[str, float],
    target_metadata: Mapping[str, TargetMember],
    execution_prices: Mapping[str, float],
) -> RebalanceResult:
    """按给定成交价格调仓，并完整返回实际买卖额和成本。"""

    missing_held_prices = sorted(set(positions) - set(execution_prices))
    if missing_held_prices:
        # 旧持仓缺少当日成交价时不能可靠卖出，因此整次调仓跳过。
        return RebalanceResult(
            positions=positions,
            cash=cash,
            pre_trade_nav=0.0,
            buy_amount=0.0,
            sell_amount=0.0,
            transaction_cost=0.0,
            trade_details=(),
            succeeded=False,
            skip_reason="旧持仓缺少当日成交价：" + ",".join(missing_held_prices),
        )

    current_values = {
        code: position.shares * execution_prices[code]
        for code, position in positions.items()
    }
    nav_before_cost = cash + sum(current_values.values())
    if nav_before_cost <= 0:
        raise RuntimeError("调仓前组合净值不大于0")

    fee = 0.0
    buy_amount = 0.0
    sell_amount = 0.0
    desired_values: dict[str, float] = {}
    for _ in range(100):
        investable_nav = max(0.0, nav_before_cost - fee)
        desired_values = {
            code: weight * investable_nav
            for code, weight in target_weights.items()
            if weight > 0
        }
        all_codes = set(current_values).union(desired_values)
        buy_amount = sum(
            max(desired_values.get(code, 0.0) - current_values.get(code, 0.0), 0.0)
            for code in all_codes
        )
        sell_amount = sum(
            max(current_values.get(code, 0.0) - desired_values.get(code, 0.0), 0.0)
            for code in all_codes
        )
        new_fee = TRANSACTION_COST_RATE * (buy_amount + sell_amount)
        if math.isclose(new_fee, fee, rel_tol=1e-13, abs_tol=1e-15):
            fee = new_fee
            break
        fee = new_fee
    else:
        raise RuntimeError("交易成本迭代未收敛")

    new_positions: dict[str, Position] = {}
    for code, desired_value in desired_values.items():
        if desired_value <= 1e-15:
            continue
        member = target_metadata[code]
        new_positions[code] = Position(
            shares=desired_value / execution_prices[code],
            index_code=member.index_code,
            index_name=member.index_name,
            etf_name=member.etf_name,
            signal_date=member.signal_date,
            trend_factor=member.trend_factor,
            factor_rank=member.factor_rank,
            selection_volume=member.selection_volume,
            target_weight=target_weights[code],
        )

    post_cost_nav = nav_before_cost - fee
    new_cash = post_cost_nav - sum(desired_values.values())
    if new_cash < 0 and abs(new_cash) <= 1e-12:
        new_cash = 0.0
    if new_cash < 0:
        raise RuntimeError(f"调仓后现金为负：{new_cash}")

    trade_details: list[TradeDetail] = []
    for code in sorted(set(current_values).union(desired_values)):
        before_value = current_values.get(code, 0.0)
        target_value = desired_values.get(code, 0.0)
        difference = target_value - before_value
        if abs(difference) <= 1e-15:
            continue
        target_member = target_metadata.get(code)
        old_position = positions.get(code)
        etf_name = (
            target_member.etf_name
            if target_member is not None
            else old_position.etf_name if old_position is not None else ""
        )
        trade_details.append(
            TradeDetail(
                etf_code=code,
                etf_name=etf_name,
                direction="买入" if difference > 0 else "卖出",
                before_value=before_value,
                target_value=target_value,
                trade_amount=abs(difference),
                transaction_cost=TRANSACTION_COST_RATE * abs(difference),
            )
        )
    return RebalanceResult(
        positions=new_positions,
        cash=new_cash,
        pre_trade_nav=nav_before_cost,
        buy_amount=buy_amount,
        sell_amount=sell_amount,
        transaction_cost=fee,
        trade_details=tuple(trade_details),
        succeeded=True,
        skip_reason="",
    )


def portfolio_value_at_close(
    positions: Mapping[str, Position],
    cash: float,
    close_prices: Mapping[str, float],
    last_closes: Mapping[str, float],
) -> float:
    value = cash
    for code, position in positions.items():
        close = close_prices.get(code, last_closes.get(code))
        if close is None:
            raise RuntimeError(f"持仓ETF缺少可用收盘价：{code}")
        value += position.shares * close
    return value


def run_backtest(
    mode: str,
    trading_dates: Sequence[date],
    targets: Mapping[date, DailyTarget],
    prices: Mapping[date, Mapping[str, PricePoint]],
) -> tuple[
    list[HoldingRecord],
    list[NavRecord],
    list[AccountDailyRecord],
    list[AccountHoldingRecord],
    list[AccountTradeRecord],
]:
    if mode not in {"close", "next_day_vwap"}:
        raise ValueError(f"未知回测模式：{mode}")
    if not trading_dates:
        raise ValueError("没有可用交易日期")

    accounts = [
        AccountState(
            account_id=account_index + 1,
            positions={},
            cash=INITIAL_NAV / ACCOUNT_COUNT,
        )
        for account_index in range(ACCOUNT_COUNT)
    ]
    for account_index, account in enumerate(accounts):
        first_position = account_index if mode == "close" else account_index + 1
        account.next_rebalance_date = (
            trading_dates[first_position]
            if first_position < len(trading_dates)
            else None
        )

    last_closes: dict[str, float] = {}
    previous_nav = INITIAL_NAV
    running_peak = INITIAL_NAV
    cumulative_cost_rate = 0.0
    holdings: list[HoldingRecord] = []
    nav_records: list[NavRecord] = []
    account_daily_records: list[AccountDailyRecord] = []
    account_holding_records: list[AccountHoldingRecord] = []
    account_trade_records: list[AccountTradeRecord] = []

    for date_position, current_date in enumerate(trading_dates):
        daily_prices = prices.get(current_date, {})
        close_prices = {
            code: point.close
            for code, point in daily_prices.items()
            if point.close is not None
        }
        vwap_prices = {
            code: point.vwap
            for code, point in daily_prices.items()
            if point.vwap is not None
        }
        last_closes.update(close_prices)

        execution_prices: Mapping[str, float] = (
            close_prices if mode == "close" else vwap_prices
        )
        scheduled_accounts = [
            account
            for account in accounts
            if account.next_rebalance_date == current_date
        ]
        if len(scheduled_accounts) > 1:
            raise RuntimeError(f"{current_date}存在多个计划调仓账户")
        rebalance_account = scheduled_accounts[0] if scheduled_accounts else None
        rebalance_attempted = rebalance_account is not None
        execution_target: DailyTarget | None = None
        signal_date: date | None = None
        rebalance_result: RebalanceResult | None = None
        missing_price_index_count = 0
        initial_build = False

        if rebalance_attempted:
            if rebalance_account is None:
                raise RuntimeError("计划调仓账户为空")
            signal_date = (
                current_date
                if mode == "close"
                else trading_dates[date_position - 1]
            )
            execution_target = targets.get(signal_date)

            if execution_target is None:
                target_weights: dict[str, float] = {}
                target_metadata: dict[str, TargetMember] = {}
            else:
                target_weights, target_metadata = aggregate_target_weights(
                    execution_target,
                    execution_prices,
                    close_prices,
                )
                missing_price_index_count = sum(
                    member.etf_code not in execution_prices
                    or member.etf_code not in close_prices
                    for member in execution_target.members
                )

            rebalance_result = rebalance_portfolio(
                rebalance_account.positions,
                rebalance_account.cash,
                target_weights,
                target_metadata,
                execution_prices,
            )
            rebalance_account.positions = rebalance_result.positions
            rebalance_account.cash = rebalance_result.cash
            next_position = date_position + ACCOUNT_REBALANCE_INTERVAL
            rebalance_account.next_rebalance_date = (
                trading_dates[next_position]
                if next_position < len(trading_dates)
                else None
            )

            if rebalance_result.succeeded:
                initial_build = (
                    not rebalance_account.has_been_built
                    and rebalance_result.buy_amount > 0
                )
                if initial_build:
                    rebalance_account.has_been_built = True
                rebalance_account.last_signal_date = signal_date
                rebalance_account.last_rebalance_date = current_date
                for detail in rebalance_result.trade_details:
                    account_trade_records.append(
                        AccountTradeRecord(
                            execution_date=current_date,
                            account_id=rebalance_account.account_id,
                            signal_date=signal_date,
                            etf_code=detail.etf_code,
                            etf_name=detail.etf_name,
                            direction=detail.direction,
                            before_value=detail.before_value,
                            target_value=detail.target_value,
                            trade_amount=detail.trade_amount,
                            transaction_cost=detail.transaction_cost,
                        )
                    )

        buy_amount = rebalance_result.buy_amount if rebalance_result else 0.0
        sell_amount = rebalance_result.sell_amount if rebalance_result else 0.0
        transaction_cost = (
            rebalance_result.transaction_cost if rebalance_result else 0.0
        )
        pre_trade_nav = previous_nav
        buy_ratio = buy_amount / pre_trade_nav if pre_trade_nav > 0 else 0.0
        sell_ratio = sell_amount / pre_trade_nav if pre_trade_nav > 0 else 0.0
        bilateral_ratio = buy_ratio + sell_ratio
        one_way_turnover = bilateral_ratio / 2.0
        transaction_cost_rate = (
            transaction_cost / pre_trade_nav if pre_trade_nav > 0 else 0.0
        )
        cumulative_cost_rate = 1.0 - (
            (1.0 - cumulative_cost_rate) * (1.0 - transaction_cost_rate)
        )

        account_navs: dict[int, float] = {}
        account_market_values: dict[int, float] = {}
        combined_market_values: dict[str, float] = defaultdict(float)
        combined_positions: dict[str, Position] = {}
        for account in accounts:
            account_market_value = 0.0
            for code, position in account.positions.items():
                close = close_prices.get(code, last_closes.get(code))
                if close is None:
                    raise RuntimeError(f"持仓ETF缺少可用收盘价：{code}")
                market_value = position.shares * close
                account_market_value += market_value
                combined_market_values[code] += market_value
                existing = combined_positions.get(code)
                if existing is None or position.signal_date > existing.signal_date:
                    combined_positions[code] = position
            account_market_values[account.account_id] = account_market_value
            account_navs[account.account_id] = account.cash + account_market_value

        nav = sum(account_navs.values())
        total_cash = sum(account.cash for account in accounts)
        daily_return = nav / previous_nav - 1.0
        gross_return = daily_return + transaction_cost / previous_nav
        running_peak = max(running_peak, nav)
        drawdown = 1.0 - nav / running_peak
        cash_weight = total_cash / nav if nav > 0 else 0.0

        capacity_samples: list[float] = []
        for code, market_value in combined_market_values.items():
            amount = daily_prices.get(code).amount if code in daily_prices else None
            actual_weight = market_value / nav if nav > 0 else 0.0
            if amount is None or actual_weight <= 1e-6:
                continue
            capacity_samples.append(
                CAPACITY_DAILY_AMOUNT_RATIO * amount / actual_weight
            )
        capacity = linear_quantile(
            capacity_samples,
            1.0 - CAPACITY_DESCENDING_QUANTILE,
        )

        nav_records.append(
            NavRecord(
                current_date=current_date,
                signal_date=signal_date,
                pre_trade_nav=pre_trade_nav,
                nav=nav,
                gross_return=gross_return,
                daily_return=daily_return,
                drawdown=drawdown,
                buy_ratio=buy_ratio,
                sell_ratio=sell_ratio,
                bilateral_ratio=bilateral_ratio,
                one_way_turnover=one_way_turnover,
                transaction_cost=transaction_cost,
                transaction_cost_rate=transaction_cost_rate,
                cumulative_cost_rate=cumulative_cost_rate,
                capacity=capacity,
                holding_count=len(combined_market_values),
                cash_weight=cash_weight,
                initial_build=initial_build,
                rebalance_account_id=(
                    rebalance_account.account_id
                    if rebalance_account is not None
                    else None
                ),
                rebalance_attempted=rebalance_attempted,
                rebalance_succeeded=(
                    rebalance_result.succeeded if rebalance_result else False
                ),
                skip_reason=(rebalance_result.skip_reason if rebalance_result else ""),
                selected_index_count=(
                    execution_target.selected_index_count if execution_target else 0
                ),
                unmapped_index_count=(
                    execution_target.unmapped_index_count if execution_target else 0
                ),
                missing_price_index_count=missing_price_index_count,
            )
        )

        for code in sorted(combined_market_values):
            position = combined_positions[code]
            actual_weight = combined_market_values[code] / nav if nav > 0 else 0.0
            holdings.append(
                HoldingRecord(
                    current_date=current_date,
                    signal_date=position.signal_date,
                    index_code=position.index_code,
                    index_name=position.index_name,
                    trend_factor=position.trend_factor,
                    factor_rank=position.factor_rank,
                    etf_code=code,
                    etf_name=position.etf_name,
                    selection_volume=position.selection_volume,
                    target_weight=actual_weight,
                    actual_weight=actual_weight,
                )
            )

        for account in accounts:
            account_nav = account_navs[account.account_id]
            account_market_value = account_market_values[account.account_id]
            account_daily_records.append(
                AccountDailyRecord(
                    current_date=current_date,
                    account_id=account.account_id,
                    account_nav=account_nav,
                    etf_market_value=account_market_value,
                    cash=account.cash,
                    cash_weight=(account.cash / account_nav if account_nav > 0 else 0.0),
                    rebalance_attempted=(account is rebalance_account),
                    rebalance_succeeded=(
                        account is rebalance_account
                        and rebalance_result is not None
                        and rebalance_result.succeeded
                    ),
                    signal_date=account.last_signal_date,
                    last_rebalance_date=account.last_rebalance_date,
                    next_rebalance_date=account.next_rebalance_date,
                )
            )
            for code, position in sorted(account.positions.items()):
                close = close_prices.get(code, last_closes.get(code))
                if close is None:
                    continue
                market_value = position.shares * close
                account_holding_records.append(
                    AccountHoldingRecord(
                        current_date=current_date,
                        account_id=account.account_id,
                        signal_date=position.signal_date,
                        etf_code=code,
                        etf_name=position.etf_name,
                        market_value=market_value,
                        account_weight=(
                            market_value / account_nav if account_nav > 0 else 0.0
                        ),
                        total_portfolio_weight=(
                            market_value / nav if nav > 0 else 0.0
                        ),
                    )
                )

        previous_nav = nav

    return (
        holdings,
        nav_records,
        account_daily_records,
        account_holding_records,
        account_trade_records,
    )


def calculate_performance(nav_records: Sequence[NavRecord]) -> dict[str, object]:
    if not nav_records:
        raise ValueError("没有净值记录，无法计算绩效")
    periods = len(nav_records)
    daily_returns = [record.daily_return for record in nav_records]
    cumulative_growth = math.prod(1.0 + value for value in daily_returns)
    cumulative_return = cumulative_growth - 1.0
    annual_return = (
        cumulative_growth ** (ANNUAL_TRADING_DAYS / periods) - 1.0
        if cumulative_growth > 0
        else float("nan")
    )
    daily_std = statistics.stdev(daily_returns) if len(daily_returns) >= 2 else 0.0
    annual_volatility = daily_std * math.sqrt(ANNUAL_TRADING_DAYS)
    daily_risk_free_rate = (
        (1.0 + ANNUAL_RISK_FREE_RATE) ** (1.0 / ANNUAL_TRADING_DAYS) - 1.0
    )
    excess_returns = [value - daily_risk_free_rate for value in daily_returns]
    sharpe = (
        statistics.fmean(excess_returns)
        / daily_std
        * math.sqrt(ANNUAL_TRADING_DAYS)
        if daily_std > 0
        else 0.0
    )
    downside_deviation = math.sqrt(
        statistics.fmean(min(value, 0.0) ** 2 for value in excess_returns)
    )
    sortino = (
        statistics.fmean(excess_returns)
        / downside_deviation
        * math.sqrt(ANNUAL_TRADING_DAYS)
        if downside_deviation > 0
        else 0.0
    )

    local_nav = 1.0
    running_peak = 1.0
    peak_date = nav_records[0].current_date
    max_drawdown = 0.0
    max_drawdown_start = peak_date
    max_drawdown_end = peak_date
    for record in nav_records:
        local_nav *= 1.0 + record.daily_return
        if local_nav >= running_peak:
            running_peak = local_nav
            peak_date = record.current_date
        drawdown = 1.0 - local_nav / running_peak
        if drawdown > max_drawdown:
            max_drawdown = drawdown
            max_drawdown_start = peak_date
            max_drawdown_end = record.current_date

    turnover_records = [record for record in nav_records if not record.initial_build]
    average_turnover = (
        statistics.fmean(record.one_way_turnover for record in turnover_records)
        if turnover_records
        else 0.0
    )
    years = periods / ANNUAL_TRADING_DAYS
    annual_turnover = (
        sum(record.one_way_turnover for record in turnover_records) / years
        if years > 0
        else 0.0
    )
    cumulative_cost_rate = 1.0 - math.prod(
        1.0 - record.transaction_cost_rate for record in nav_records
    )

    return {
        "回测开始日": nav_records[0].current_date.isoformat(),
        "回测结束日": nav_records[-1].current_date.isoformat(),
        "交易日数量": periods,
        "初始净值": INITIAL_NAV,
        "期末净值": nav_records[-1].nav,
        "累计收益率": cumulative_return,
        "年化收益率": annual_return,
        "年化波动率": annual_volatility,
        "夏普比率": sharpe,
        "Sortino比率": sortino,
        "最大回撤": max_drawdown,
        "最大回撤开始日": max_drawdown_start.isoformat(),
        "最大回撤结束日": max_drawdown_end.isoformat(),
        "Calmar比率": annual_return / max_drawdown if max_drawdown > 0 else 0.0,
        "平均每日单边换手率": average_turnover,
        "年化单边换手率（倍）": annual_turnover,
        "累计交易成本率": cumulative_cost_rate,
        "平均持仓ETF数量": statistics.fmean(
            record.holding_count for record in nav_records
        ),
        "调仓次数": sum(record.rebalance_succeeded for record in nav_records),
        "跳过调仓次数": sum(
            record.rebalance_attempted and not record.rebalance_succeeded
            for record in nav_records
        ),
        "无法映射指数-日期记录数": sum(
            record.unmapped_index_count for record in nav_records
        ),
        "缺少成交价指数-日期记录数": sum(
            record.missing_price_index_count for record in nav_records
        ),
    }


MODE_LABELS = {
    "close": "收盘价成交",
    "next_day_vwap": "次日VWAP成交",
}


def build_annual_metrics(
    mode: str,
    nav_records: Sequence[NavRecord],
) -> list[dict[str, object]]:
    by_year: dict[int, list[NavRecord]] = defaultdict(list)
    for record in nav_records:
        by_year[record.current_date.year].append(record)

    rows: list[dict[str, object]] = []
    for year in sorted(by_year):
        records = by_year[year]
        performance = calculate_performance(records)
        start = records[0].current_date
        end = records[-1].current_date
        if start.month > 1:
            year_label = f"{year}年{start:%m-%d}起"
        elif end.month < 12:
            year_label = f"{year}年截至{end:%m-%d}"
        else:
            year_label = f"{year}年"
        rows.append(
            {
                "成交方式": MODE_LABELS[mode],
                "年份": year_label,
                "交易日数量": performance["交易日数量"],
                "年度收益率": performance["累计收益率"],
                "年化收益率": performance["年化收益率"],
                "年化波动率": performance["年化波动率"],
                "夏普比率": performance["夏普比率"],
                "Sortino比率": performance["Sortino比率"],
                "最大回撤": performance["最大回撤"],
                "最大回撤开始日": performance["最大回撤开始日"],
                "最大回撤结束日": performance["最大回撤结束日"],
                "Calmar比率": performance["Calmar比率"],
                "平均每日单边换手率": performance["平均每日单边换手率"],
                "年化单边换手率（倍）": performance["年化单边换手率（倍）"],
                "累计交易成本率": performance["累计交易成本率"],
                "平均持仓ETF数量": performance["平均持仓ETF数量"],
            }
        )
    return rows


def build_validation_rows(
    mode: str,
    holding_records: Sequence[HoldingRecord],
    nav_records: Sequence[NavRecord],
) -> list[dict[str, object]]:
    label = MODE_LABELS[mode]
    holdings_by_date: dict[date, float] = defaultdict(float)
    for record in holding_records:
        holdings_by_date[record.current_date] += record.actual_weight

    nav_rebuild_error = 0.0
    previous_nav = INITIAL_NAV
    weight_error = 0.0
    cost_error = 0.0
    signal_errors = 0
    finite_errors = 0
    for record in nav_records:
        rebuilt_nav = previous_nav * (1.0 + record.daily_return)
        nav_rebuild_error = max(nav_rebuild_error, abs(rebuilt_nav - record.nav))
        previous_nav = record.nav
        weight_error = max(
            weight_error,
            abs(holdings_by_date.get(record.current_date, 0.0) + record.cash_weight - 1.0),
        )
        cost_error = max(
            cost_error,
            abs(
                record.transaction_cost_rate
                - TRANSACTION_COST_RATE * record.bilateral_ratio
            ),
        )
        if record.signal_date is not None:
            if mode == "close" and record.signal_date != record.current_date:
                signal_errors += 1
            if mode == "next_day_vwap" and record.signal_date >= record.current_date:
                signal_errors += 1
        numeric_values = (
            record.nav,
            record.daily_return,
            record.one_way_turnover,
            record.transaction_cost_rate,
            record.cash_weight,
        )
        finite_errors += sum(not math.isfinite(value) for value in numeric_values)

    initial_build_accounts = [
        record.rebalance_account_id
        for record in nav_records
        if record.initial_build and record.rebalance_account_id is not None
    ]
    initial_build_count = len(initial_build_accounts)
    initial_build_valid = (
        len(initial_build_accounts) == len(set(initial_build_accounts))
        and initial_build_count <= ACCOUNT_COUNT
    )
    skipped_count = sum(
        record.rebalance_attempted and not record.rebalance_succeeded
        for record in nav_records
    )

    def check_row(
        item: str,
        actual: object,
        expected: object,
        difference: float,
        tolerance: float,
        passed: bool,
        notes: str,
    ) -> dict[str, object]:
        return {
            "成交方式": label,
            "检查项": item,
            "实际值": actual,
            "期望值": expected,
            "差异": difference,
            "容差": tolerance,
            "状态": "OK" if passed else "FAIL",
            "说明": notes,
        }

    return [
        check_row(
            "净值可由日收益重建",
            nav_rebuild_error,
            0.0,
            nav_rebuild_error,
            1e-12,
            nav_rebuild_error <= 1e-12,
            "逐日用前一日净值×(1+净收益率)重建",
        ),
        check_row(
            "持仓权重与现金权重合计为1",
            weight_error,
            0.0,
            weight_error,
            2e-10,
            weight_error <= 2e-10,
            "使用每日收盘估值后的实际权重",
        ),
        check_row(
            "交易成本率与双边成交额一致",
            cost_error,
            0.0,
            cost_error,
            1e-12,
            cost_error <= 1e-12,
            "成本率=0.1%×(买入比例+卖出比例)",
        ),
        check_row(
            "信号日期与成交日期无未来数据",
            signal_errors,
            0,
            float(signal_errors),
            0.0,
            signal_errors == 0,
            "收盘成交同日；次日VWAP成交必须晚于信号日",
        ),
        check_row(
            "关键日序列均为有限数值",
            finite_errors,
            0,
            float(finite_errors),
            0.0,
            finite_errors == 0,
            "净值、收益、换手、成本和现金权重",
        ),
        check_row(
            "各账户初始建仓单独标记",
            initial_build_count,
            "每个账户至多1次",
            0.0 if initial_build_valid else 1.0,
            0.0,
            initial_build_valid,
            f"{ACCOUNT_COUNT}个账户分别至多初始建仓1次；收费，但不计入平均和年化换手率",
        ),
        check_row(
            "跳过调仓次数（信息项）",
            skipped_count,
            "仅记录",
            0.0,
            0.0,
            True,
            "信息项：记录回测中未成功执行的调仓次数",
        ),
    ]


def style_worksheet(
    sheet: Any,
    percent_headers: set[str] | None = None,
    decimal_headers: set[str] | None = None,
    integer_headers: set[str] | None = None,
    max_width: int = 28,
) -> None:
    percent_headers = percent_headers or set()
    decimal_headers = decimal_headers or set()
    integer_headers = integer_headers or set()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    body_font = Font(name="Arial", size=10)
    thin_gray = Side(style="thin", color="D9D9D9")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
        )
    sheet.row_dimensions[1].height = 24
    headers = {cell.value: cell.column for cell in sheet[1]}
    for header in percent_headers:
        column = headers.get(header)
        if column:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = "0.00%"
    for header in decimal_headers:
        column = headers.get(header)
        if column:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = "0.000000"
    for header in integer_headers:
        column = headers.get(header)
        if column:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = "#,##0"
    for column_number in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_number)
        values = [sheet.cell(row=row, column=column_number).value for row in range(1, min(sheet.max_row, 300) + 1)]
        width = min(
            max((len(str(value)) for value in values if value is not None), default=10) + 2,
            max_width,
        )
        sheet.column_dimensions[column_letter].width = max(width, 11)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = Border(
                left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
            )
            cell.alignment = Alignment(
                horizontal="center" if isinstance(cell.value, str) else "right",
                vertical="center",
            )


def save_figure_atomic(figure: Any, path: Path) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    try:
        figure.savefig(
            temporary,
            format="png",
            dpi=180,
            bbox_inches="tight",
            facecolor="white",
        )
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)
        plt.close(figure)



def save_workbook_atomic(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)
        workbook.close()
    return path


def mode_execution_label(mode: str) -> str:
    return "收盘价" if mode == "close" else "次日VWAP"


def write_annual_metrics_workbook(
    mode: str,
    nav_records: Sequence[NavRecord],
    output_dir: Path,
) -> Path:
    """按参考项目单独输出年度指标表，不包含基准和超额字段。"""

    headers = [
        "年份",
        "策略收益",
        "年化波动",
        "Sharpe",
        "Sortino",
        "策略最大回撤",
        "Calmar",
        "年化换手率（单边）",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "annual_metrics"
    sheet.append(headers)
    for row in build_annual_metrics(mode, nav_records):
        sheet.append(
            [
                row["年份"],
                row["年度收益率"],
                row["年化波动率"],
                row["夏普比率"],
                row["Sortino比率"],
                row["最大回撤"],
                row["Calmar比率"],
                row["年化单边换手率（倍）"],
            ]
        )
    style_worksheet(
        sheet,
        percent_headers={"策略收益", "年化波动", "策略最大回撤"},
        decimal_headers={"Sharpe", "Sortino", "Calmar"},
        max_width=24,
    )
    turnover_column = headers.index("年化换手率（单边）") + 1
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=turnover_column).number_format = '0.00"倍"'
    return save_workbook_atomic(
        workbook,
        output_dir / f"{mode}_annual_metrics.xlsx",
    )


def write_backtest_metrics_workbook(
    mode: str,
    nav_records: Sequence[NavRecord],
    output_dir: Path,
    score_method: str,
) -> Path:
    """按参考项目单独输出总回测指标和参数。"""

    performance = calculate_performance(nav_records)
    attempted = sum(record.rebalance_attempted for record in nav_records)
    succeeded = sum(record.rebalance_succeeded for record in nav_records)
    success_rate = succeeded / attempted if attempted else 0.0
    drawdown_range = (
        f"{datetime.fromisoformat(str(performance['最大回撤开始日'])).strftime('%Y年%m月')}"
        f"-{datetime.fromisoformat(str(performance['最大回撤结束日'])).strftime('%Y年%m月')}"
    )

    workbook = Workbook()
    performance_sheet = workbook.active
    performance_sheet.title = "performance"
    performance_sheet.append(["类别", "指标", "ETF趋势策略"])
    performance_rows = [
        ["基本信息", "回测区间", f"{performance['回测开始日']}至{performance['回测结束日']}"],
        ["基本信息", "交易日数量", performance["交易日数量"]],
        ["基本信息", "初始净值", performance["初始净值"]],
        ["基本信息", "期末净值", performance["期末净值"]],
        ["基本信息", "累计收益率", performance["累计收益率"]],
        ["收益表现", "策略年化收益率", performance["年化收益率"]],
        ["绝对风险收益", "策略年化波动率", performance["年化波动率"]],
        ["绝对风险收益", "Sharpe", performance["夏普比率"]],
        ["绝对风险收益", "Sortino", performance["Sortino比率"]],
        ["绝对风险收益", "策略最大回撤", performance["最大回撤"]],
        ["绝对风险收益", "策略最大回撤区间", drawdown_range],
        ["绝对风险收益", "Calmar", performance["Calmar比率"]],
        ["交易与组合", "年化换手率（单边）", performance["年化单边换手率（倍）"]],
        ["交易与组合", "平均单次换仓比率（单边）", performance["平均每日单边换手率"]],
        ["交易与组合", "累计交易成本率", performance["累计交易成本率"]],
        ["交易与组合", "平均持仓ETF数量", performance["平均持仓ETF数量"]],
        ["交易与组合", "调仓次数", performance["调仓次数"]],
        ["交易与组合", "调仓成功率", success_rate],
    ]
    for row in performance_rows:
        performance_sheet.append(row)
    style_worksheet(performance_sheet, max_width=30)
    percentage_metrics = {
        "累计收益率",
        "策略年化收益率",
        "策略年化波动率",
        "策略最大回撤",
        "平均单次换仓比率（单边）",
        "累计交易成本率",
        "调仓成功率",
    }
    for row_number in range(2, performance_sheet.max_row + 1):
        metric = performance_sheet.cell(row=row_number, column=2).value
        value_cell = performance_sheet.cell(row=row_number, column=3)
        if metric in percentage_metrics:
            value_cell.number_format = "0.00%"
        elif metric == "年化换手率（单边）":
            value_cell.number_format = '0.00"倍"'
        elif metric == "交易日数量":
            value_cell.number_format = '0"天"'
        elif metric == "调仓次数":
            value_cell.number_format = '0"次"'
        elif metric == "平均持仓ETF数量":
            value_cell.number_format = '0.00"只"'
        elif metric in {"初始净值", "期末净值", "Sharpe", "Sortino", "Calmar"}:
            value_cell.number_format = "0.0000"
    performance_sheet.column_dimensions["A"].width = 18
    performance_sheet.column_dimensions["B"].width = 34
    performance_sheet.column_dimensions["C"].width = 28

    parameter_sheet = workbook.create_sheet("parameters")
    parameter_sheet.append(["类别", "参数 / 约束", "本项目设置"])
    execution_label = mode_execution_label(mode)
    if REBALANCE_MODE == "rebalance":
        account_structure_label = (
            f"单账户每{ACCOUNT_REBALANCE_INTERVAL}个交易日全组合调仓"
        )
        planned_rebalance_label = "计划调仓日1个，其他交易日0个"
        transfer_label = "不适用；单账户"
    else:
        account_structure_label = (
            f"{ACCOUNT_COUNT}个独立账户逐日错峰持有"
            f"{ACCOUNT_REBALANCE_INTERVAL}个交易日"
        )
        planned_rebalance_label = "每个交易日1个"
        transfer_label = "无；各账户独立复利并保留自己的现金"
    parameter_rows = [
        ["基本信息", "执行价格", execution_label],
        ["趋势策略", "聚类相关性阈值", CLUSTER_CORRELATION_THRESHOLD],
        ["趋势策略", "趋势因子窗口", TREND_WINDOW],
        ["趋势策略", "排名得分公式", SCORE_LABELS[score_method]],
        ["趋势策略", "调仓日入选比例", TOP_PERCENT],
        ["趋势过滤", "过滤位置", "排名后"],
        ["趋势过滤", "过滤条件", "当前趋势窗口收益率>0"],
        ["趋势过滤", "未通过处理", "不向后补选，空缺权重留现金"],
        ["ETF选择", "代表ETF选择", "跟踪同一指数中当日成交量最大"],
        ["账户结构", "调仓模式", REBALANCE_MODE],
        ["账户结构", "组合结构", account_structure_label],
        ["账户结构", "账户数量", ACCOUNT_COUNT],
        ["账户结构", "每账户初始资金比例", 1.0 / ACCOUNT_COUNT],
        ["账户结构", "计划调仓账户数", planned_rebalance_label],
        ["账户结构", "单账户调仓间隔（交易日）", ACCOUNT_REBALANCE_INTERVAL],
        ["账户结构", "账户之间资金转移", transfer_label],
        ["交易设置", "初始净值 NAV₀", INITIAL_NAV],
        ["收益参数", "年化无风险利率 r_f", ANNUAL_RISK_FREE_RATE],
        ["交易成本", "买入成本 c_buy", TRANSACTION_COST_RATE],
        ["交易成本", "卖出成本 c_sell", TRANSACTION_COST_RATE],
        ["交易成本", "完整换仓成本 T_c", 2.0 * TRANSACTION_COST_RATE],
        ["容量参数", "当日成交额使用比例 ρ_amt", CAPACITY_DAILY_AMOUNT_RATIO],
        ["容量参数", "倒序分位 q_desc", CAPACITY_DESCENDING_QUANTILE],
        ["容量参数", "等价升序分位 q_asc", 1.0 - CAPACITY_DESCENDING_QUANTILE],
        ["组合约束", "指数权重", "按过滤前计划入选数量等权"],
        ["收益口径", "扣费后收益", "买入和卖出均扣除交易成本"],
        ["容量口径", "组合容量", "各持仓ETF容量的5%分位"],
    ]
    for row in parameter_rows:
        parameter_sheet.append(row)
    style_worksheet(parameter_sheet, max_width=50)
    percentage_parameters = {
        "调仓日入选比例",
        "每账户初始资金比例",
        "年化无风险利率 r_f",
        "买入成本 c_buy",
        "卖出成本 c_sell",
        "完整换仓成本 T_c",
        "当日成交额使用比例 ρ_amt",
        "倒序分位 q_desc",
        "等价升序分位 q_asc",
    }
    for row_number in range(2, parameter_sheet.max_row + 1):
        parameter = parameter_sheet.cell(row=row_number, column=2).value
        if parameter in percentage_parameters:
            parameter_sheet.cell(row=row_number, column=3).number_format = "0.00%"
    parameter_sheet.column_dimensions["A"].width = 18
    parameter_sheet.column_dimensions["B"].width = 40
    parameter_sheet.column_dimensions["C"].width = 52

    return save_workbook_atomic(
        workbook,
        output_dir / f"{mode}_backtest_metrics.xlsx",
    )


def write_holdings_workbook(
    mode: str,
    holding_records: Sequence[HoldingRecord],
    output_dir: Path,
) -> Path:
    """按参考项目单独输出每日持仓。"""

    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "holdings"
    sheet.append(["日期", "ETF代码", "ETF名称", "权重"])
    for record in holding_records:
        sheet.append(
            [
                int(record.current_date.strftime("%Y%m%d")),
                record.etf_code,
                record.etf_name,
                record.actual_weight,
            ]
        )
    style_worksheet(sheet, decimal_headers={"权重"}, max_width=28)
    weight_column = 4
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=weight_column).number_format = "0.000000"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 38
    sheet.column_dimensions["D"].width = 16
    return save_workbook_atomic(
        workbook,
        output_dir / f"{mode}_holdings.xlsx",
    )


def write_time_series_workbook(
    mode: str,
    nav_records: Sequence[NavRecord],
    output_dir: Path,
) -> Path:
    """按参考项目单独输出策略时序，不包含基准和超额字段。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "time_series"
    headers = [
        "日期",
        "策略日收益",
        "策略累计净值",
        "每日换仓比率（单边）",
        "策略可容纳规模（亿元）",
    ]
    sheet.append(headers)
    for record in nav_records:
        sheet.append(
            [
                record.current_date,
                record.daily_return,
                record.nav,
                record.one_way_turnover,
                record.capacity / 1e8 if record.capacity is not None else None,
            ]
        )
    style_worksheet(
        sheet,
        percent_headers={"策略日收益", "每日换仓比率（单边）"},
        decimal_headers={"策略累计净值", "策略可容纳规模（亿元）"},
        max_width=34,
    )
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=row_number, column=3).number_format = "0.000000"
        sheet.cell(row=row_number, column=5).number_format = "0.0000"
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 30
    return save_workbook_atomic(
        workbook,
        output_dir / f"{mode}_time_series.xlsx",
    )


def write_account_details_workbook(
    mode: str,
    daily_records: Sequence[AccountDailyRecord],
    holding_records: Sequence[AccountHoldingRecord],
    trade_records: Sequence[AccountTradeRecord],
    output_dir: Path,
) -> Path:
    """输出各独立账户的每日状态、账户持仓和实际交易差额。"""

    workbook = Workbook()
    daily_sheet = workbook.active
    daily_sheet.title = "账户每日状态"
    daily_headers = [
        "日期", "账户编号", "账户净值", "ETF市值", "现金", "现金权重",
        "当天是否调仓", "调仓是否成功", "当前信号日", "上次调仓日", "下次计划调仓日",
    ]
    daily_sheet.append(daily_headers)
    for record in daily_records:
        daily_sheet.append(
            [
                record.current_date,
                record.account_id,
                record.account_nav,
                record.etf_market_value,
                record.cash,
                record.cash_weight,
                "是" if record.rebalance_attempted else "否",
                "是" if record.rebalance_succeeded else "否",
                record.signal_date,
                record.last_rebalance_date,
                record.next_rebalance_date,
            ]
        )
    style_worksheet(
        daily_sheet,
        percent_headers={"现金权重"},
        decimal_headers={"账户净值", "ETF市值", "现金"},
        integer_headers={"账户编号"},
        max_width=24,
    )
    for row_number in range(2, daily_sheet.max_row + 1):
        for column_number in (1, 9, 10, 11):
            daily_sheet.cell(row=row_number, column=column_number).number_format = (
                "yyyy-mm-dd"
            )

    holding_sheet = workbook.create_sheet("账户持仓")
    holding_headers = [
        "日期", "账户编号", "信号日", "ETF代码", "ETF名称", "ETF市值",
        "账户内部权重", "对总组合贡献权重",
    ]
    holding_sheet.append(holding_headers)
    for record in holding_records:
        holding_sheet.append(
            [
                record.current_date,
                record.account_id,
                record.signal_date,
                record.etf_code,
                record.etf_name,
                record.market_value,
                record.account_weight,
                record.total_portfolio_weight,
            ]
        )
    style_worksheet(
        holding_sheet,
        percent_headers={"账户内部权重", "对总组合贡献权重"},
        decimal_headers={"ETF市值"},
        integer_headers={"账户编号"},
        max_width=32,
    )
    for row_number in range(2, holding_sheet.max_row + 1):
        holding_sheet.cell(row=row_number, column=1).number_format = "yyyy-mm-dd"
        holding_sheet.cell(row=row_number, column=3).number_format = "yyyy-mm-dd"

    trade_sheet = workbook.create_sheet("账户交易")
    trade_headers = [
        "成交日期", "账户编号", "信号日", "ETF代码", "ETF名称", "交易方向",
        "交易前市值", "最新目标市值", "实际成交金额", "交易成本",
    ]
    trade_sheet.append(trade_headers)
    for record in trade_records:
        trade_sheet.append(
            [
                record.execution_date,
                record.account_id,
                record.signal_date,
                record.etf_code,
                record.etf_name,
                record.direction,
                record.before_value,
                record.target_value,
                record.trade_amount,
                record.transaction_cost,
            ]
        )
    style_worksheet(
        trade_sheet,
        decimal_headers={"交易前市值", "最新目标市值", "实际成交金额", "交易成本"},
        integer_headers={"账户编号"},
        max_width=32,
    )
    for row_number in range(2, trade_sheet.max_row + 1):
        trade_sheet.cell(row=row_number, column=1).number_format = "yyyy-mm-dd"
        trade_sheet.cell(row=row_number, column=3).number_format = "yyyy-mm-dd"

    return save_workbook_atomic(
        workbook,
        output_dir / f"{mode}_account_details.xlsx",
    )


def write_mode_workbooks(
    mode: str,
    holding_records: Sequence[HoldingRecord],
    nav_records: Sequence[NavRecord],
    account_daily_records: Sequence[AccountDailyRecord],
    account_holding_records: Sequence[AccountHoldingRecord],
    account_trade_records: Sequence[AccountTradeRecord],
    score_method: str,
    score_backtest_dir: Path,
) -> list[Path]:
    output_dir = score_backtest_dir / mode
    output_dir.mkdir(parents=True, exist_ok=True)
    return [
        write_annual_metrics_workbook(mode, nav_records, output_dir),
        write_backtest_metrics_workbook(
            mode,
            nav_records,
            output_dir,
            score_method,
        ),
        write_holdings_workbook(mode, holding_records, output_dir),
        write_time_series_workbook(mode, nav_records, output_dir),
        write_account_details_workbook(
            mode,
            account_daily_records,
            account_holding_records,
            account_trade_records,
            output_dir,
        ),
    ]


def style_plot_axis(axis: Any) -> None:
    axis.grid(axis="y", color="#D9D9D9", linewidth=0.6)
    axis.spines[["top", "right"]].set_visible(False)


def write_mode_figures(
    mode: str,
    nav_records: Sequence[NavRecord],
    score_backtest_dir: Path,
) -> list[Path]:
    """每种交易模式独立生成图表，不和另一模式叠加。"""

    output_dir = score_backtest_dir / mode
    output_dir.mkdir(parents=True, exist_ok=True)
    plt.style.use("default")
    plt.rcParams["font.sans-serif"] = [
        "Arial Unicode MS", "PingFang SC", "Heiti SC", "SimHei", "DejaVu Sans"
    ]
    plt.rcParams["axes.unicode_minus"] = False
    color = "#1F4E79" if mode == "close" else "#C0504D"
    label = mode_execution_label(mode)
    dates = [record.current_date for record in nav_records]

    nav_path = output_dir / f"{mode}_cumulative_nav.png"
    figure, axis = plt.subplots(figsize=(10.0, 4.8), facecolor="white")
    axis.plot(
        dates,
        [record.nav for record in nav_records],
        color=color,
        linewidth=1.5,
        label="ETF趋势策略",
    )
    axis.set_title(
        f"Plot 1: Cumulative Net Value ({label})",
        loc="left",
        fontsize=13,
        fontweight="bold",
    )
    axis.legend(frameon=False)
    axis.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=5, maxticks=9))
    axis.xaxis.set_major_formatter(
        mdates.ConciseDateFormatter(axis.xaxis.get_major_locator())
    )
    style_plot_axis(axis)
    save_figure_atomic(figure, nav_path)

    turnover_path = output_dir / f"{mode}_turnover.png"
    figure, axis = plt.subplots(figsize=(10.0, 4.5), facecolor="white")
    axis.bar(
        range(len(nav_records)),
        [record.one_way_turnover for record in nav_records],
        width=1.0,
        color=color,
        label="组合单次换仓比率（单边）",
    )
    axis.set_title(
        f"Plot 3: Rebalance Turnover ({label})",
        loc="left",
        fontsize=13,
        fontweight="bold",
    )
    axis.set_xlabel("调仓序号")
    axis.yaxis.set_major_formatter(PercentFormatter(1.0))
    axis.legend(frameon=False)
    style_plot_axis(axis)
    save_figure_atomic(figure, turnover_path)

    cost_path = output_dir / f"{mode}_transaction_cost.png"
    figure, axis = plt.subplots(figsize=(10.0, 4.5), facecolor="white")
    axis.plot(
        dates,
        [record.cumulative_cost_rate for record in nav_records],
        color=color,
        linewidth=1.5,
        label="累计交易成本率",
    )
    axis.set_title(
        f"Cumulative Transaction Cost Rate ({label})",
        loc="left",
        fontsize=13,
        fontweight="bold",
    )
    axis.yaxis.set_major_formatter(PercentFormatter(1.0))
    axis.legend(frameon=False)
    axis.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=5, maxticks=9))
    axis.xaxis.set_major_formatter(
        mdates.ConciseDateFormatter(axis.xaxis.get_major_locator())
    )
    style_plot_axis(axis)
    save_figure_atomic(figure, cost_path)

    capacity_path = output_dir / f"{mode}_capacity.png"
    capacity_dates = [
        record.current_date for record in nav_records if record.capacity is not None
    ]
    capacity_values = [
        record.capacity / 1e8
        for record in nav_records
        if record.capacity is not None
    ]
    figure, axis = plt.subplots(figsize=(10.0, 4.5), facecolor="white")
    axis.plot(
        capacity_dates,
        capacity_values,
        color=color,
        linewidth=1.3,
        label="ETF趋势策略",
    )
    axis.set_title(
        f"Plot 4: Strategy Capacity ({label})",
        loc="left",
        fontsize=13,
        fontweight="bold",
    )
    axis.set_ylabel("亿元")
    axis.legend(frameon=False)
    axis.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=5, maxticks=9))
    axis.xaxis.set_major_formatter(
        mdates.ConciseDateFormatter(axis.xaxis.get_major_locator())
    )
    style_plot_axis(axis)
    save_figure_atomic(figure, capacity_path)
    return [nav_path, turnover_path, cost_path, capacity_path]


def validate_mode_outputs(
    mode: str,
    holding_records: Sequence[HoldingRecord],
    nav_records: Sequence[NavRecord],
    account_daily_records: Sequence[AccountDailyRecord],
    account_holding_records: Sequence[AccountHoldingRecord],
    account_trade_records: Sequence[AccountTradeRecord],
    workbook_paths: Sequence[Path],
    figure_paths: Sequence[Path],
) -> None:
    failures = [
        row
        for row in build_validation_rows(mode, holding_records, nav_records)
        if row["状态"] == "FAIL"
    ]
    if failures:
        raise RuntimeError(
            f"{mode}回测内部校验失败："
            + "、".join(str(row["检查项"]) for row in failures)
        )

    nav_by_date = {record.current_date: record for record in nav_records}
    account_rows_by_date: dict[date, list[AccountDailyRecord]] = defaultdict(list)
    for record in account_daily_records:
        account_rows_by_date[record.current_date].append(record)
    first_rebalance_position = 0 if mode == "close" else 1
    for date_position, nav_record in enumerate(nav_records):
        current_date = nav_record.current_date
        rows = account_rows_by_date.get(current_date, [])
        if len(rows) != ACCOUNT_COUNT:
            raise RuntimeError(
                f"{mode}账户校验失败：{current_date}不是{ACCOUNT_COUNT}个账户"
            )
        account_nav_sum = sum(row.account_nav for row in rows)
        if not math.isclose(account_nav_sum, nav_record.nav, abs_tol=2e-12):
            raise RuntimeError(f"{mode}账户校验失败：{current_date}账户净值合计不一致")
        attempted_count = sum(row.rebalance_attempted for row in rows)
        if REBALANCE_MODE == "rebalance":
            expected_attempted = int(
                date_position >= first_rebalance_position
                and (
                    date_position - first_rebalance_position
                ) % ACCOUNT_REBALANCE_INTERVAL == 0
            )
        else:
            expected_attempted = int(date_position >= first_rebalance_position)
        if attempted_count != expected_attempted:
            raise RuntimeError(f"{mode}账户校验失败：{current_date}调仓账户数不正确")

    date_positions = {
        record.current_date: position for position, record in enumerate(nav_records)
    }
    attempted_dates_by_account: dict[int, list[date]] = defaultdict(list)
    for record in account_daily_records:
        if record.rebalance_attempted:
            attempted_dates_by_account[record.account_id].append(record.current_date)
    for account_id, attempted_dates in attempted_dates_by_account.items():
        gaps = [
            date_positions[later] - date_positions[earlier]
            for earlier, later in zip(attempted_dates, attempted_dates[1:])
        ]
        if any(gap != ACCOUNT_REBALANCE_INTERVAL for gap in gaps):
            raise RuntimeError(
                f"{mode}账户校验失败：账户{account_id}并非每"
                f"{ACCOUNT_REBALANCE_INTERVAL}个交易日调仓"
            )

    trade_cost_by_date: dict[date, float] = defaultdict(float)
    for record in account_trade_records:
        trade_cost_by_date[record.execution_date] += record.transaction_cost
    for current_date, nav_record in nav_by_date.items():
        if not math.isclose(
            trade_cost_by_date.get(current_date, 0.0),
            nav_record.transaction_cost,
            abs_tol=2e-12,
        ):
            raise RuntimeError(f"{mode}账户校验失败：{current_date}交易成本不一致")

    expected_workbooks = {
        f"{mode}_annual_metrics.xlsx": ["annual_metrics"],
        f"{mode}_backtest_metrics.xlsx": ["performance", "parameters"],
        f"{mode}_holdings.xlsx": ["holdings"],
        f"{mode}_time_series.xlsx": ["time_series"],
        f"{mode}_account_details.xlsx": ["账户每日状态", "账户持仓", "账户交易"],
    }
    expected_headers = {
        (f"{mode}_annual_metrics.xlsx", "annual_metrics"): [
            "年份", "策略收益", "年化波动", "Sharpe", "Sortino",
            "策略最大回撤", "Calmar", "年化换手率（单边）",
        ],
        (f"{mode}_backtest_metrics.xlsx", "performance"): [
            "类别", "指标", "ETF趋势策略",
        ],
        (f"{mode}_backtest_metrics.xlsx", "parameters"): [
            "类别", "参数 / 约束", "本项目设置",
        ],
        (f"{mode}_holdings.xlsx", "holdings"): [
            "日期", "ETF代码", "ETF名称", "权重",
        ],
        (f"{mode}_time_series.xlsx", "time_series"): [
            "日期", "策略日收益", "策略累计净值",
            "每日换仓比率（单边）", "策略可容纳规模（亿元）",
        ],
        (f"{mode}_account_details.xlsx", "账户每日状态"): [
            "日期", "账户编号", "账户净值", "ETF市值", "现金", "现金权重",
            "当天是否调仓", "调仓是否成功", "当前信号日", "上次调仓日", "下次计划调仓日",
        ],
        (f"{mode}_account_details.xlsx", "账户持仓"): [
            "日期", "账户编号", "信号日", "ETF代码", "ETF名称", "ETF市值",
            "账户内部权重", "对总组合贡献权重",
        ],
        (f"{mode}_account_details.xlsx", "账户交易"): [
            "成交日期", "账户编号", "信号日", "ETF代码", "ETF名称", "交易方向",
            "交易前市值", "最新目标市值", "实际成交金额", "交易成本",
        ],
    }
    actual_workbooks = {path.name: path for path in workbook_paths}
    if set(actual_workbooks) != set(expected_workbooks):
        raise RuntimeError(f"{mode}回测Excel文件结构不正确")
    for filename, expected_sheets in expected_workbooks.items():
        path = actual_workbooks[filename]
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if workbook.sheetnames != expected_sheets:
                raise RuntimeError(
                    f"{filename}的sheet不正确：{workbook.sheetnames}"
                )
            for sheet in workbook.worksheets:
                allow_header_only = (
                    filename == f"{mode}_account_details.xlsx"
                    and sheet.title in {"账户持仓", "账户交易"}
                )
                if (sheet.max_row < 2 and not allow_header_only) or sheet.max_column < 1:
                    raise RuntimeError(f"{filename}/{sheet.title}没有有效数据")
                actual_headers = [
                    cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
                ]
                if actual_headers != expected_headers[(filename, sheet.title)]:
                    raise RuntimeError(
                        f"{filename}/{sheet.title}的列结构不正确"
                    )
        finally:
            workbook.close()

    for path in [*workbook_paths, *figure_paths]:
        if not path.exists() or path.stat().st_size == 0:
            raise RuntimeError(f"输出文件缺失或为空：{path}")


def main() -> None:
    validate_parameters()
    if REBALANCE_MODE == "rebalance":
        rebalance_description = (
            f"全组合每 {ACCOUNT_REBALANCE_INTERVAL} 个交易日调仓一次"
        )
    else:
        rebalance_description = (
            f"初始资金分成 {ACCOUNT_COUNT} 个独立账户，每天轮换1个账户，"
            f"每账户持有 {ACCOUNT_REBALANCE_INTERVAL} 个交易日"
        )
    print(
        f"聚类阈值 {CLUSTER_CORRELATION_THRESHOLD:g}，"
        f"趋势窗口 {TREND_WINDOW}，选择调仓信号日得分前 {TOP_PERCENT:.0%}；"
        f"排名后过滤窗口收益率不大于 {MIN_WINDOW_RETURN:g} 的指数，"
        f"{rebalance_description}；"
        "本次依次回测两种得分公式。",
        flush=True,
    )
    for score_method in SCORE_METHODS_TO_RUN:
        score_column = SCORE_COLUMNS[score_method]
        score_label = SCORE_LABELS[score_method]
        score_backtest_dir = (
            BACKTEST_DIR
            / score_method
            / STRATEGY_VARIANT_DIR
            / ACCOUNT_VARIANT_DIR
        )
        print(f"\n开始回测：{score_label}（{score_column}）", flush=True)

        daily_selections = read_daily_top_factors(score_column)
        targets, full_trading_calendar = build_daily_targets(daily_selections)
        first_signal_date = min(targets)
        last_signal_date = max(targets)
        trading_dates = [
            current_date
            for current_date in full_trading_calendar
            if first_signal_date <= current_date <= last_signal_date
        ]
        missing_calendar_dates = sorted(set(targets) - set(trading_dates))
        if missing_calendar_dates:
            raise ValueError(
                "趋势信号日期不在ETF交易日历中："
                + ",".join(value.isoformat() for value in missing_calendar_dates[:10])
            )
        selected_codes = {
            member.etf_code
            for target in targets.values()
            for member in target.members
        }
        unmapped_count = sum(
            target.unmapped_index_count for target in targets.values()
        )
        filtered_count = sum(
            target.filtered_index_count for target in targets.values()
        )
        print(
            f"共 {len(trading_dates)} 个实际ETF交易日，"
            f"实际涉及 {len(selected_codes)} 只ETF，"
            f"排名后过滤指数-日期记录 {filtered_count} 条，"
            f"无法映射的指数-日期记录 {unmapped_count} 条。",
            flush=True,
        )
        prices = load_selected_prices(selected_codes, set(trading_dates))

        for mode in ("close", "next_day_vwap"):
            (
                holdings,
                nav_records,
                account_daily_records,
                account_holding_records,
                account_trade_records,
            ) = run_backtest(
                mode,
                trading_dates,
                targets,
                prices,
            )
            workbook_paths = write_mode_workbooks(
                mode,
                holdings,
                nav_records,
                account_daily_records,
                account_holding_records,
                account_trade_records,
                score_method,
                score_backtest_dir,
            )
            figure_paths = write_mode_figures(
                mode,
                nav_records,
                score_backtest_dir,
            )
            validate_mode_outputs(
                mode,
                holdings,
                nav_records,
                account_daily_records,
                account_holding_records,
                account_trade_records,
                workbook_paths,
                figure_paths,
            )
            performance = calculate_performance(nav_records)
            print(
                f"✅ {score_label} / {mode} 回测完成：年化收益率 "
                f"{float(performance['年化收益率']):.2%}，"
                f"Sharpe {float(performance['夏普比率']):.2f}，"
                f"最大回撤 {float(performance['最大回撤']):.2%}，"
                f"平均每日单边换手率 "
                f"{float(performance['平均每日单边换手率']):.2%}；"
                f"输出目录：{score_backtest_dir / mode}\n"
                + "\n".join(
                    f"  {path.name}"
                    for path in [*workbook_paths, *figure_paths]
                ),
                flush=True,
            )


if __name__ == "__main__":
    main()

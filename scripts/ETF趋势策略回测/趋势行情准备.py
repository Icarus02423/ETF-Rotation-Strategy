#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
为各期动态指数池准备趋势因子所需的连续指数行情。

输入：
- outputs/etf_pool/clusters/threshold_<相关性阈值>/reports/YYYY_MM_DD.xlsx
  只读取第一张“动态指数池”。

输出：
- outputs/etf_trend_strategy/threshold_<相关性阈值>/index_prices/
  YYYY_MM_DD.csv

输出CSV的列、内容和生成规则与原聚类脚本中的趋势行情输出完全一致。
"""

from __future__ import annotations

import csv
import importlib.util
import math
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from types import ModuleType
from typing import Mapping, Sequence

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]

# ============================= 行情参数 =============================
# 必须与要读取的聚类结果阈值一致，可改为0.7、0.8或0.9。
CLUSTER_CORRELATION_THRESHOLD = 0.9

TREND_HISTORY_LOOKBACK_CALENDAR_DAYS = 180
TREND_MAX_LOOKBACK_CALENDAR_DAYS = 730
TREND_REQUIRED_HISTORY_CLOSES = 61

# False：完整的月度输出直接跳过，避免重复调用历史行情接口。
# True：忽略已有月度输出并重新下载、覆盖全部月份。
OVERWRITE_EXISTING_MONTHS = False
# ====================================================================

ALLOWED_CLUSTER_THRESHOLDS = (0.7, 0.8, 0.9)
CLUSTER_POOL_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_pool"
    / "clusters"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "reports"
)
OUTPUT_DIR = (
    PROJECT_ROOT
    / "outputs"
    / "etf_trend_strategy"
    / f"threshold_{CLUSTER_CORRELATION_THRESHOLD:g}"
    / "index_prices"
)
INDEX_DOWNLOADER_SCRIPT = (
    PROJECT_ROOT / "scripts" / "ETF池筛选" / "指数收益率准备.py"
)

POOL_SHEET_NAME = "动态指数池"
POOL_REQUIRED_COLUMNS = {"日期", "对标指数", "对标指数代码"}
OUTPUT_COLUMNS = [
    "月末交易日",
    "收益日期",
    "对标指数代码",
    "对标指数",
    "收盘价",
    "日收益率",
]


@dataclass(frozen=True)
class TrendPoolInput:
    """一期动态指数池及其实际生效行情区间。"""

    selection_date: date
    end_date: date
    index_names: Mapping[str, str]

    @property
    def output_file(self) -> Path:
        return OUTPUT_DIR / self.selection_date.strftime("%Y_%m_%d.csv")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "--", "None", "null", "NULL"} else text


def parse_date(value: object, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    for date_format in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name}不是有效日期：{text!r}")


def parse_finite_number(value: object) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def validate_parameters() -> None:
    if not any(
        math.isclose(
            CLUSTER_CORRELATION_THRESHOLD,
            allowed,
            abs_tol=1e-12,
        )
        for allowed in ALLOWED_CLUSTER_THRESHOLDS
    ):
        raise ValueError(
            "CLUSTER_CORRELATION_THRESHOLD只能设为0.7、0.8或0.9"
        )
    if TREND_REQUIRED_HISTORY_CLOSES < max(20, 40, 60):
        raise ValueError("TREND_REQUIRED_HISTORY_CLOSES不能少于60")
    if TREND_HISTORY_LOOKBACK_CALENDAR_DAYS <= TREND_REQUIRED_HISTORY_CLOSES:
        raise ValueError("趋势行情初始回溯自然日设置过短")
    if TREND_MAX_LOOKBACK_CALENDAR_DAYS < TREND_HISTORY_LOOKBACK_CALENDAR_DAYS:
        raise ValueError("趋势行情最大回溯天数不能小于初始回溯天数")


def discover_trend_pool_inputs() -> list[TrendPoolInput]:
    if not CLUSTER_POOL_DIR.exists():
        raise FileNotFoundError(f"找不到月底动态指数池：{CLUSTER_POOL_DIR}")

    files = sorted(CLUSTER_POOL_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"动态指数池目录没有XLSX：{CLUSTER_POOL_DIR}")

    snapshots: list[tuple[date, dict[str, str]]] = []
    for path in files:
        try:
            file_date = datetime.strptime(path.stem, "%Y_%m_%d").date()
        except ValueError as exc:
            raise ValueError(f"动态指数池文件名日期无效：{path.name}") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if POOL_SHEET_NAME not in workbook.sheetnames:
                raise ValueError(f"{path.name}缺少工作表：{POOL_SHEET_NAME}")
            sheet = workbook[POOL_SHEET_NAME]
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise ValueError(f"{path.name}的{POOL_SHEET_NAME}为空") from exc
            headers = [clean_text(value) for value in header_row]
            missing_columns = POOL_REQUIRED_COLUMNS - set(headers)
            if missing_columns:
                raise ValueError(
                    f"{path.name}的{POOL_SHEET_NAME}缺少列："
                    f"{sorted(missing_columns)}"
                )
            position = {header: index for index, header in enumerate(headers)}

            index_names: dict[str, str] = {}
            for row_number, values in enumerate(rows, start=2):
                if not any(value is not None for value in values):
                    continue
                row_date = parse_date(values[position["日期"]], "日期")
                if row_date != file_date:
                    raise ValueError(
                        f"{path.name}第{row_number}行日期{row_date}与文件名不一致"
                    )
                index_code = clean_text(values[position["对标指数代码"]])
                index_name = clean_text(values[position["对标指数"]])
                if not index_code:
                    raise ValueError(f"{path.name}第{row_number}行指数代码为空")
                if index_code in index_names:
                    raise ValueError(f"{path.name}存在重复代表指数：{index_code}")
                index_names[index_code] = index_name
        finally:
            workbook.close()

        if not index_names:
            raise ValueError(f"{path.name}的{POOL_SHEET_NAME}没有指数")
        snapshots.append((file_date, dict(sorted(index_names.items()))))

    selection_dates = [selection_date for selection_date, _ in snapshots]
    if len(selection_dates) != len(set(selection_dates)):
        raise ValueError("动态指数池存在重复月份")

    pools: list[TrendPoolInput] = []
    for position, (selection_date, index_names) in enumerate(snapshots):
        end_date = (
            snapshots[position + 1][0]
            if position + 1 < len(snapshots)
            else selection_date
        )
        pools.append(
            TrendPoolInput(
                selection_date=selection_date,
                end_date=end_date,
                index_names=index_names,
            )
        )
    return pools


def load_index_downloader_module() -> ModuleType:
    """加载指数收益率准备脚本，复用同一套历史行情接口与回退配置。"""

    if not INDEX_DOWNLOADER_SCRIPT.exists():
        raise FileNotFoundError(f"找不到指数收益率准备脚本：{INDEX_DOWNLOADER_SCRIPT}")
    module_name = "index_return_preparation_for_trend_data"
    spec = importlib.util.spec_from_file_location(module_name, INDEX_DOWNLOADER_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载指数收益率准备脚本：{INDEX_DOWNLOADER_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)

    required_names = (
        "HISTORY_ENDPOINT",
        "AKSHARE_FALLBACK_CONFIG",
        "load_etf_downloader_module",
        "merge_windows",
        "uncovered_windows",
        "request_index_close_range",
        "request_akshare_fallback_closes",
    )
    missing_names = [name for name in required_names if not hasattr(module, name)]
    if missing_names:
        raise RuntimeError(
            f"指数收益率准备.py缺少必要配置或函数：{'、'.join(missing_names)}"
        )
    return module


def output_is_complete(pool: TrendPoolInput) -> bool:
    path = pool.output_file
    if not path.exists():
        return False

    dates_by_code: dict[str, set[date]] = defaultdict(set)
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if list(reader.fieldnames or []) != OUTPUT_COLUMNS:
                return False
            for row in reader:
                if parse_date(row.get("月末交易日"), "月末交易日") != pool.selection_date:
                    return False
                index_code = clean_text(row.get("对标指数代码"))
                if index_code not in pool.index_names:
                    return False
                price_date = parse_date(row.get("收益日期"), "收益日期")
                if price_date > pool.end_date or price_date in dates_by_code[index_code]:
                    return False
                close = parse_finite_number(row.get("收盘价"))
                daily_return = parse_finite_number(row.get("日收益率"))
                if close is None or close <= 0 or daily_return is None:
                    return False
                dates_by_code[index_code].add(price_date)
    except (OSError, csv.Error, ValueError):
        return False

    if set(dates_by_code) != set(pool.index_names):
        return False
    for index_code in pool.index_names:
        price_dates = dates_by_code[index_code]
        history_count = sum(price_date <= pool.selection_date for price_date in price_dates)
        if history_count < TREND_REQUIRED_HISTORY_CLOSES:
            return False
        if pool.end_date > pool.selection_date and not any(
            pool.selection_date < price_date <= pool.end_date
            for price_date in price_dates
        ):
            return False
    return True


def requirements_without_enough_history(
    pools: Sequence[TrendPoolInput],
    closes_by_code: Mapping[str, Mapping[date, float]],
) -> list[tuple[TrendPoolInput, str]]:
    required_raw_closes = TREND_REQUIRED_HISTORY_CLOSES + 1
    insufficient: list[tuple[TrendPoolInput, str]] = []
    for pool in pools:
        for index_code in pool.index_names:
            history_count = sum(
                price_date <= pool.selection_date
                for price_date in closes_by_code.get(index_code, {})
            )
            if history_count < required_raw_closes:
                insufficient.append((pool, index_code))
    return insufficient


def build_output_rows(
    pool: TrendPoolInput,
    closes_by_code: Mapping[str, Mapping[date, float]],
) -> tuple[list[list[str]], list[str]]:
    required_raw_closes = TREND_REQUIRED_HISTORY_CLOSES + 1
    rows: list[list[str]] = []
    insufficient_codes: list[str] = []

    for index_code in sorted(pool.index_names):
        available = closes_by_code.get(index_code, {})
        history = sorted(
            (price_date, close)
            for price_date, close in available.items()
            if price_date <= pool.selection_date
        )
        future = sorted(
            (price_date, close)
            for price_date, close in available.items()
            if pool.selection_date < price_date <= pool.end_date
        )
        if len(history) < required_raw_closes:
            insufficient_codes.append(index_code)
            continue
        if pool.end_date > pool.selection_date and not future:
            insufficient_codes.append(index_code)
            continue

        selected = history[-required_raw_closes:] + future
        for previous, current in zip(selected, selected[1:]):
            price_date, close = current
            daily_return = close / previous[1] - 1.0
            rows.append(
                [
                    pool.selection_date.isoformat(),
                    price_date.isoformat(),
                    index_code,
                    pool.index_names[index_code],
                    format(close, ".15g"),
                    format(daily_return, ".15g"),
                ]
            )
    return rows, sorted(set(insufficient_codes))


def write_output(path: Path, rows: Sequence[Sequence[str]]) -> None:
    temp_path = path.with_name(f".{path.name}.tmp")
    try:
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(OUTPUT_COLUMNS)
            writer.writerows(rows)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def download_trend_index_data(pools: Sequence[TrendPoolInput]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pending_pools = [
        pool
        for pool in pools
        if OVERWRITE_EXISTING_MONTHS or not output_is_complete(pool)
    ]
    for pool in pools:
        if pool not in pending_pools:
            print(
                f"⏭️ {pool.output_file.name} 的代表指数连续行情已存在，跳过下载",
                flush=True,
            )
    if not pending_pools:
        print(f"全部代表指数连续行情已存在：{OUTPUT_DIR}", flush=True)
        return

    index_module = load_index_downloader_module()
    downloader = index_module.load_etf_downloader_module()
    refresh_token = clean_text(downloader.IFIND_REFRESH_TOKEN)
    if not refresh_token:
        raise ValueError("下载ETF数据.py中的IFIND_REFRESH_TOKEN为空")

    print("正在登录同花顺 iFinD，准备下载代表指数连续行情...", flush=True)
    access_token = downloader.get_access_token(refresh_token, "iFinD账号")

    windows_by_code: dict[str, list[tuple[date, date]]] = defaultdict(list)
    pools_by_code: dict[str, list[TrendPoolInput]] = defaultdict(list)
    for pool in pending_pools:
        start_date = pool.selection_date - timedelta(
            days=TREND_HISTORY_LOOKBACK_CALENDAR_DAYS
        )
        for index_code in pool.index_names:
            windows_by_code[index_code].append((start_date, pool.end_date))
            pools_by_code[index_code].append(pool)

    closes_by_code: dict[str, dict[date, float]] = defaultdict(dict)
    requested_windows: dict[str, list[tuple[date, date]]] = defaultdict(list)
    errors: dict[str, str] = {}
    codes = sorted(windows_by_code)

    def apply_akshare_fallback(index_codes: Sequence[str]) -> None:
        for index_code in index_codes:
            related_pools = pools_by_code[index_code]
            fallback_start = min(
                pool.selection_date for pool in related_pools
            ) - timedelta(days=TREND_MAX_LOOKBACK_CALENDAR_DAYS)
            fallback_end = max(pool.end_date for pool in related_pools)
            api_name, symbol, _, _ = index_module.AKSHARE_FALLBACK_CONFIG[index_code]
            print(
                f"⚠️ {index_code} iFinD失败，改用AKShare "
                f"{api_name}({symbol})读取同一指数...",
                flush=True,
            )
            try:
                closes_by_code[index_code].update(
                    index_module.request_akshare_fallback_closes(
                        index_code,
                        fallback_start,
                        fallback_end,
                    )
                )
                requested_windows[index_code].append((fallback_start, fallback_end))
                errors.pop(index_code, None)
            except Exception as exc:
                errors[index_code] = f"AKShare回退失败：{exc}"
                print(f"❌ {index_code} AKShare回退失败：{exc}", flush=True)

    for position, index_code in enumerate(codes, start=1):
        windows = index_module.merge_windows(windows_by_code[index_code])
        print(
            f"正在下载 {index_code} 代表指数连续收盘价... "
            f"({position}/{len(codes)})",
            flush=True,
        )
        try:
            for start_date, end_date in windows:
                requested_windows[index_code].append((start_date, end_date))
                closes_by_code[index_code].update(
                    index_module.request_index_close_range(
                        downloader,
                        access_token,
                        index_code,
                        start_date,
                        end_date,
                    )
                )
            print(
                f"✅ {index_code} 连续行情下载完成，"
                f"取得 {len(closes_by_code[index_code])} 个收盘价",
                flush=True,
            )
        except Exception as exc:
            errors[index_code] = str(exc)
            print(f"❌ {index_code} 连续行情下载失败：{exc}", flush=True)

    apply_akshare_fallback(
        sorted(set(errors) & set(index_module.AKSHARE_FALLBACK_CONFIG))
    )

    insufficient_requirements = requirements_without_enough_history(
        pending_pools,
        closes_by_code,
    )
    blocked_pool_dates = {
        pool.selection_date
        for pool in pending_pools
        if any(index_code in errors for index_code in pool.index_names)
    }
    extra_windows_by_code: dict[str, list[tuple[date, date]]] = defaultdict(list)
    for pool, index_code in insufficient_requirements:
        if pool.selection_date in blocked_pool_dates or index_code in errors:
            continue
        extra_windows_by_code[index_code].extend(
            index_module.uncovered_windows(
                pool.selection_date
                - timedelta(days=TREND_MAX_LOOKBACK_CALENDAR_DAYS),
                pool.selection_date,
                requested_windows[index_code],
            )
        )

    for index_code in sorted(extra_windows_by_code):
        windows = index_module.merge_windows(extra_windows_by_code[index_code])
        if not windows:
            continue
        print(f"正在向前补充 {index_code} 代表指数历史收盘价...", flush=True)
        try:
            for start_date, end_date in windows:
                requested_windows[index_code].append((start_date, end_date))
                closes_by_code[index_code].update(
                    index_module.request_index_close_range(
                        downloader,
                        access_token,
                        index_code,
                        start_date,
                        end_date,
                    )
                )
        except Exception as exc:
            errors[index_code] = str(exc)
            print(f"❌ {index_code} 历史收盘价补充失败：{exc}", flush=True)

    apply_akshare_fallback(
        sorted(set(errors) & set(index_module.AKSHARE_FALLBACK_CONFIG))
    )

    failed_files: dict[str, list[str]] = {}
    written_count = 0
    for pool in pending_pools:
        rows, insufficient_codes = build_output_rows(pool, closes_by_code)
        failed_codes = sorted(
            set(insufficient_codes).union(
                index_code
                for index_code in pool.index_names
                if index_code in errors
            )
        )
        if failed_codes:
            failed_files[pool.output_file.name] = failed_codes
            print(
                f"❌ {pool.output_file.name} 未保存代表指数连续行情："
                f"{'、'.join(failed_codes)}",
                flush=True,
            )
            continue
        write_output(pool.output_file, rows)
        written_count += 1
        print(
            f"✅ 已保存 {pool.output_file.name} 代表指数连续行情："
            f"{pool.output_file}",
            flush=True,
        )

    print(
        f"代表指数连续行情完成：本次写入 {written_count} 个月度CSV，"
        f"输出目录：{OUTPUT_DIR}",
        flush=True,
    )
    if failed_files:
        raise RuntimeError(
            f"仍有 {len(failed_files)} 个月度代表指数连续行情未保存；"
            "查看上方日志中的指数代码。"
        )


def main() -> None:
    validate_parameters()
    pools = discover_trend_pool_inputs()
    print(
        f"聚类阈值 {CLUSTER_CORRELATION_THRESHOLD:g}，"
        f"共 {len(pools)} 期动态指数池。",
        flush=True,
    )
    download_trend_index_data(pools)


if __name__ == "__main__":
    main()

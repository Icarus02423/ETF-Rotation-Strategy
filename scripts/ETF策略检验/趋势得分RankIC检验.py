"""全面检验两种ETF趋势得分的预测能力、可交易性与稳定性。

核心口径：
1. 因子在信号日收盘后形成，下一ETF交易日按VWAP买入，持有1至20个交易日后
   按VWAP卖出；ETF收益同时输出扣费前和按双边0.1%成本估算的扣费后结果。
2. 全池Rank IC用于判断因子整体单调性；正收益池、前20%和前10%局部Rank IC
   用于观察策略候选区间，但小样本日期会明确标记而不是强行计算。
3. 同时使用ETF的VWAP未来收益和对标指数的收盘未来收益。前者贴近交易，后者
   用于隔离代表ETF选择、溢折价和跟踪误差带来的噪声。
4. 除Rank IC外，输出Q1至Q5分组收益、Q5-Q1、多头候选组合、年度与滚动稳定性、
   Newey-West显著性、数据覆盖率和异常样本。

脚本仍按得分公式分别生成原有两个工作簿，并在同一输出目录生成配套图片：
outputs/etf_strategy_test/trend_rank_ic/01_收益率乘R平方.xlsx
outputs/etf_strategy_test/trend_rank_ic/02_收益率除以波动率.xlsx
"""

from __future__ import annotations

import csv
import heapq
import math
from bisect import bisect_right
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Sequence

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.ticker import PercentFormatter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
FACTOR_ROOT = PROJECT_ROOT / "outputs" / "etf_trend_strategy"
ETF_DATA_FILE = PROJECT_ROOT / "outputs" / "etf_data" / "etf_data.csv"
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "etf_strategy_test" / "trend_rank_ic"

HOLDING_DAYS = tuple(range(1, 21))
MIN_WINDOW_RETURN = 0.0
LOCAL_SCOPE_PERCENTAGES = {
    "前20%策略池": 0.20,
    "前10%策略池": 0.10,
}
MIN_CROSS_SECTION = 5
QUANTILE_COUNT = 5
ROLLING_IC_WINDOW = 60
ROLLING_HOLDING_DAYS = (1, 5, 10, 20)
TRANSACTION_COST_RATE = 0.001
ANNUAL_TRADING_DAYS = 252
MAX_ANOMALY_ROWS = 300
EXCEL_MAX_DATA_ROWS = 1_048_575

RETURN_BASIS_ETF = "ETF次日VWAP"
RETURN_BASIS_INDEX = "对标指数收盘"
SCOPE_ALL = "全池"
SCOPE_POSITIVE = "窗口收益率>0"

SCORE_COLUMNS = {
    "收益率×R平方": "趋势质量因子",
    "收益率÷波动率": "风险调整趋势得分",
}
OUTPUT_FILES = {
    "收益率×R平方": OUTPUT_DIR / "01_收益率乘R平方.xlsx",
    "收益率÷波动率": OUTPUT_DIR / "02_收益率除以波动率.xlsx",
}
FACTOR_COLUMNS = {"日期", "对标指数代码", "窗口收益率"} | set(SCORE_COLUMNS.values())
ETF_COLUMNS = {
    "日期",
    "代码",
    "上市日期",
    "对标指数代码",
    "规模",
    "成交量",
    "成交额",
    "VWAP",
}
INDEX_PRICE_COLUMNS = {"收益日期", "对标指数代码", "收盘价"}


@dataclass(frozen=True)
class CandidateETF:
    code: str
    volume: float
    amount: float
    scale: float


@dataclass(frozen=True)
class DailyAssetReturn:
    index_code: str
    etf_code: str | None
    score: float
    window_return: float
    etf_gross_return: float | None
    etf_net_return: float | None
    index_return: float | None


@dataclass
class CoverageStats:
    signal_days: int = 0
    scheduled_days: int = 0
    factor_observations: int = 0
    scheduled_factor_observations: int = 0
    mapped_observations: int = 0
    etf_return_observations: int = 0
    index_return_observations: int = 0


@dataclass(frozen=True)
class ResearchFrames:
    parameters: pd.DataFrame
    factor_distribution: pd.DataFrame
    data_quality: pd.DataFrame
    ic_summary: pd.DataFrame
    daily_ic: pd.DataFrame
    annual_ic: pd.DataFrame
    rolling_ic: pd.DataFrame
    quantile_returns: pd.DataFrame
    long_short_returns: pd.DataFrame
    top_portfolios: pd.DataFrame
    robustness: pd.DataFrame
    anomalies: pd.DataFrame


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "null"} else text


def parse_date(value: object) -> date | None:
    text = clean_text(value)
    if not text:
        return None
    compact_date = text[:8]
    if len(compact_date) == 8 and compact_date.isdigit():
        try:
            return date(
                int(compact_date[:4]),
                int(compact_date[4:6]),
                int(compact_date[6:8]),
            )
        except ValueError:
            return None
    try:
        return date.fromisoformat(text[:10].replace("/", "-"))
    except ValueError:
        return None


def finite_float(value: object) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def positive_float(value: object) -> float | None:
    number = finite_float(value)
    return number if number is not None and number > 0 else None


def candidate_sort_key(candidate: CandidateETF) -> tuple[float, float, float, str]:
    return (-candidate.volume, -candidate.amount, -candidate.scale, candidate.code)


def parse_configuration(window_dir: Path) -> tuple[float, int]:
    try:
        threshold = float(window_dir.parents[1].name.removeprefix("threshold_"))
        trend_window = int(window_dir.name.removeprefix("window_"))
    except ValueError as exc:
        raise ValueError(f"无法识别因子目录参数：{window_dir}") from exc
    return threshold, trend_window


def load_factor_signals() -> pd.DataFrame:
    """读取所有已有参数组合的完整因子截面。"""

    if not FACTOR_ROOT.exists():
        raise FileNotFoundError(f"找不到趋势因子目录：{FACTOR_ROOT}")

    frames: list[pd.DataFrame] = []
    window_dirs = sorted(FACTOR_ROOT.glob("threshold_*/factors/window_*"))
    for window_dir in window_dirs:
        threshold, trend_window = parse_configuration(window_dir)
        for factor_file in sorted(window_dir.glob("*.csv")):
            frame = pd.read_csv(
                factor_file,
                encoding="utf-8-sig",
                usecols=lambda column: column in FACTOR_COLUMNS,
                dtype={"对标指数代码": "string"},
                low_memory=False,
            )
            missing = FACTOR_COLUMNS - set(frame.columns)
            if missing:
                raise ValueError(
                    f"{factor_file} 缺少列：{sorted(missing)}。"
                    "请先用趋势因子计算.py重新生成这个参数组合。"
                )

            frame = frame.rename(
                columns={
                    "日期": "signal_date",
                    "对标指数代码": "index_code",
                    "窗口收益率": "window_return",
                }
            )
            frame["signal_date"] = pd.to_datetime(
                frame["signal_date"], errors="coerce"
            ).dt.date
            frame["index_code"] = frame["index_code"].str.strip()
            score_columns = list(SCORE_COLUMNS.values())
            numeric_columns = score_columns + ["window_return"]
            for score_column in numeric_columns:
                frame[score_column] = pd.to_numeric(
                    frame[score_column],
                    errors="coerce",
                )
            frame = frame.dropna(subset=["window_return"])
            frame = frame[frame["window_return"].map(math.isfinite)]

            frame = frame.melt(
                id_vars=["signal_date", "index_code", "window_return"],
                value_vars=score_columns,
                var_name="score_column",
                value_name="score",
            )
            score_labels = {
                column: label for label, column in SCORE_COLUMNS.items()
            }
            frame["score_method"] = frame["score_column"].map(score_labels)
            frame = frame.drop(columns="score_column")
            frame["threshold"] = threshold
            frame["trend_window"] = trend_window
            frames.append(frame)

    if not frames:
        raise FileNotFoundError(f"没有在 {FACTOR_ROOT} 下找到因子CSV")

    signals = pd.concat(frames, ignore_index=True)
    signals = signals.dropna(subset=["signal_date", "index_code", "score"])
    signals = signals[signals["index_code"].ne("")]
    signals = signals[signals["score"].map(math.isfinite)]
    signals = signals.drop_duplicates(
        subset=[
            "threshold",
            "trend_window",
            "signal_date",
            "index_code",
            "score_method",
        ],
        keep="last",
    )
    signals = signals.sort_values(
        [
            "threshold",
            "trend_window",
            "signal_date",
            "score_method",
            "index_code",
        ]
    ).reset_index(drop=True)
    if signals.empty:
        raise ValueError("因子CSV中没有可用于Rank IC检验的有效记录")
    return signals


def scan_etf_mapping(
    signals: pd.DataFrame,
) -> tuple[dict[tuple[date, str], CandidateETF], list[date]]:
    """第一次扫描ETF总表：建立交易日历和信号日的指数到ETF映射。"""

    if not ETF_DATA_FILE.exists():
        raise FileNotFoundError(f"找不到ETF数据：{ETF_DATA_FILE}")

    required_indexes: dict[date, set[str]] = defaultdict(set)
    for row in signals[["signal_date", "index_code"]].itertuples(index=False):
        required_indexes[row.signal_date].add(row.index_code)

    trading_dates: set[date] = set()
    best_by_date_index: dict[tuple[date, str], CandidateETF] = {}

    with ETF_DATA_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = ETF_COLUMNS - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"etf_data.csv缺少列：{sorted(missing)}")

        for row in reader:
            current_date = parse_date(row.get("日期"))
            if current_date is None:
                continue
            trading_dates.add(current_date)

            index_code = clean_text(row.get("对标指数代码"))
            if index_code not in required_indexes.get(current_date, set()):
                continue

            code = clean_text(row.get("代码")).upper()
            volume = positive_float(row.get("成交量"))
            if not code or volume is None:
                continue

            listed_text = clean_text(row.get("上市日期"))
            if listed_text:
                listed_date = parse_date(listed_text)
                if listed_date is None or listed_date > current_date:
                    continue

            candidate = CandidateETF(
                code=code,
                volume=volume,
                amount=finite_float(row.get("成交额")) or 0.0,
                scale=finite_float(row.get("规模")) or 0.0,
            )
            key = (current_date, index_code)
            existing = best_by_date_index.get(key)
            if existing is None or candidate_sort_key(candidate) < candidate_sort_key(existing):
                best_by_date_index[key] = candidate

    if not trading_dates:
        raise ValueError("ETF总表中没有有效交易日期")
    return best_by_date_index, sorted(trading_dates)


def build_trade_schedule(
    signal_dates: set[date], trading_dates: list[date]
) -> dict[tuple[date, int], tuple[date, date]]:
    """生成信号日对应的下一交易日买入及1至20日后的卖出日期。"""

    schedule: dict[tuple[date, int], tuple[date, date]] = {}
    for signal_date in signal_dates:
        entry_position = bisect_right(trading_dates, signal_date)
        if entry_position >= len(trading_dates):
            continue
        entry_date = trading_dates[entry_position]
        for holding_day in HOLDING_DAYS:
            exit_position = entry_position + holding_day
            if exit_position < len(trading_dates):
                schedule[(signal_date, holding_day)] = (
                    entry_date,
                    trading_dates[exit_position],
                )
    return schedule


def collect_required_prices(
    signals: pd.DataFrame,
    mapping: dict[tuple[date, str], CandidateETF],
    schedule: dict[tuple[date, int], tuple[date, date]],
) -> set[tuple[date, str]]:
    required: set[tuple[date, str]] = set()
    for row in signals[["signal_date", "index_code"]].itertuples(index=False):
        candidate = mapping.get((row.signal_date, row.index_code))
        if candidate is None:
            continue
        for holding_day in HOLDING_DAYS:
            dates = schedule.get((row.signal_date, holding_day))
            if dates is None:
                continue
            entry_date, exit_date = dates
            required.add((entry_date, candidate.code))
            required.add((exit_date, candidate.code))
    return required


def load_required_vwap(
    required: set[tuple[date, str]],
) -> dict[tuple[date, str], float]:
    """第二次扫描ETF总表，只保留计算未来收益所需的VWAP。"""

    prices: dict[tuple[date, str], float] = {}
    if not required:
        return prices

    required_codes = {code for _, code in required}
    with ETF_DATA_FILE.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            code = clean_text(row.get("代码")).upper()
            if code not in required_codes:
                continue
            current_date = parse_date(row.get("日期"))
            key = (current_date, code) if current_date is not None else None
            if key not in required:
                continue
            vwap = positive_float(row.get("VWAP"))
            if vwap is not None:
                prices[key] = vwap
    return prices


def load_index_closes(
    signals: pd.DataFrame,
) -> dict[tuple[float, date, str], float]:
    """读取各阈值快照中的指数收盘价，用于检验因子本身而非ETF映射。"""

    required_codes = {
        float(threshold): set(group["index_code"])
        for threshold, group in signals.groupby("threshold", sort=True)
    }
    prices: dict[tuple[float, date, str], float] = {}

    for threshold, index_codes in required_codes.items():
        price_dir = FACTOR_ROOT / f"threshold_{threshold:g}" / "index_prices"
        if not price_dir.exists():
            continue
        for price_file in sorted(price_dir.glob("*.csv")):
            frame = pd.read_csv(
                price_file,
                encoding="utf-8-sig",
                usecols=lambda column: column in INDEX_PRICE_COLUMNS,
                dtype={"对标指数代码": "string"},
                low_memory=False,
            )
            missing = INDEX_PRICE_COLUMNS - set(frame.columns)
            if missing:
                raise ValueError(f"{price_file} 缺少列：{sorted(missing)}")

            frame = frame[["收益日期", "对标指数代码", "收盘价"]]
            frame["对标指数代码"] = frame["对标指数代码"].str.strip()
            frame = frame[frame["对标指数代码"].isin(index_codes)]
            for row in frame.itertuples(index=False, name=None):
                current_date = parse_date(row[0])
                index_code = clean_text(row[1])
                close = positive_float(row[2])
                if current_date is None or not index_code or close is None:
                    continue
                key = (threshold, current_date, index_code)
                existing = prices.get(key)
                if existing is not None and not math.isclose(
                    existing,
                    close,
                    rel_tol=1e-10,
                    abs_tol=1e-8,
                ):
                    raise ValueError(
                        f"指数收盘价冲突：threshold={threshold:g}, "
                        f"date={current_date}, index={index_code}, "
                        f"values=({existing}, {close})"
                    )
                prices[key] = close
    return prices


def spearman_rank_ic(
    scores: Sequence[float],
    returns: Sequence[float],
    min_count: int = MIN_CROSS_SECTION,
) -> float | None:
    if len(scores) < min_count or len(scores) != len(returns):
        return None
    score_ranks = pd.Series(scores, dtype="float64").rank(method="average")
    return_ranks = pd.Series(returns, dtype="float64").rank(method="average")
    if score_ranks.nunique() < 2 or return_ranks.nunique() < 2:
        return None
    rank_ic = score_ranks.corr(return_ranks)
    return float(rank_ic) if pd.notna(rank_ic) and math.isfinite(rank_ic) else None


def round_trip_net_return(gross_return: float) -> float:
    """按买卖双边相同费率，把毛收益换算为净收益。"""

    return (
        (1.0 + gross_return)
        * (1.0 - TRANSACTION_COST_RATE)
        / (1.0 + TRANSACTION_COST_RATE)
        - 1.0
    )


def newey_west_statistics(
    values: Iterable[float],
    lag: int,
) -> dict[str, float | int | None]:
    array = np.asarray(
        [float(value) for value in values if math.isfinite(float(value))],
        dtype="float64",
    )
    count = int(array.size)
    if count == 0:
        return {
            "样本数": 0,
            "均值": None,
            "中位数": None,
            "标准差": None,
            "ICIR": None,
            "年化ICIR": None,
            "正值比例": None,
            "大于0.02比例": None,
            "NW_t值": None,
            "NW_p值": None,
        }

    mean = float(array.mean())
    median = float(np.median(array))
    standard_deviation = float(array.std(ddof=1)) if count > 1 else None
    icir = (
        mean / standard_deviation
        if standard_deviation is not None and standard_deviation > 0
        else None
    )

    centered = array - mean
    effective_lag = min(max(int(lag), 0), count - 1)
    long_run_variance = float(np.dot(centered, centered) / count)
    for offset in range(1, effective_lag + 1):
        weight = 1.0 - offset / (effective_lag + 1.0)
        autocovariance = float(
            np.dot(centered[offset:], centered[:-offset]) / count
        )
        long_run_variance += 2.0 * weight * autocovariance
    standard_error = math.sqrt(max(long_run_variance, 0.0) / count)
    t_value = mean / standard_error if standard_error > 0 else None
    p_value = (
        math.erfc(abs(t_value) / math.sqrt(2.0))
        if t_value is not None
        else None
    )

    return {
        "样本数": count,
        "均值": mean,
        "中位数": median,
        "标准差": standard_deviation,
        "ICIR": icir,
        "年化ICIR": (
            icir * math.sqrt(ANNUAL_TRADING_DAYS) if icir is not None else None
        ),
        "正值比例": float(np.mean(array > 0)),
        "大于0.02比例": float(np.mean(array > 0.02)),
        "NW_t值": t_value,
        "NW_p值": p_value,
    }


def scope_index_codes(group: pd.DataFrame) -> dict[str, set[str]]:
    """先按分数划定候选池，再做正收益过滤；不从池外补位。"""

    ranked = group.sort_values(
        ["score", "index_code"],
        ascending=[False, True],
    )
    scopes = {
        SCOPE_ALL: set(ranked["index_code"]),
        SCOPE_POSITIVE: set(
            ranked.loc[ranked["window_return"] > MIN_WINDOW_RETURN, "index_code"]
        ),
    }
    for scope_name, percentage in LOCAL_SCOPE_PERCENTAGES.items():
        planned_count = max(1, math.ceil(len(ranked) * percentage))
        selected = ranked.head(planned_count)
        selected = selected[selected["window_return"] > MIN_WINDOW_RETURN]
        scopes[scope_name] = set(selected["index_code"])
    return scopes


def rank_ic_status(
    scores: Sequence[float],
    returns: Sequence[float],
) -> tuple[float | None, str]:
    if len(scores) < MIN_CROSS_SECTION:
        return None, f"样本不足<{MIN_CROSS_SECTION}"
    rank_ic = spearman_rank_ic(scores, returns)
    if rank_ic is None:
        return None, "得分或收益无截面差异"
    return rank_ic, "有效"


def calculate_factor_distribution(signals: pd.DataFrame) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for (threshold, trend_window), group in signals.groupby(
        ["threshold", "trend_window"], sort=True
    ):
        scores = group["score"].astype(float)
        records.append(
            {
                "聚类阈值": threshold,
                "趋势窗口": int(trend_window),
                "信号日数": int(group["signal_date"].nunique()),
                "因子观测数": int(len(group)),
                "日均截面数": float(
                    group.groupby("signal_date").size().mean()
                ),
                "得分均值": float(scores.mean()),
                "得分标准差": float(scores.std(ddof=1)),
                "最小值": float(scores.min()),
                "P01": float(scores.quantile(0.01)),
                "P05": float(scores.quantile(0.05)),
                "中位数": float(scores.median()),
                "P95": float(scores.quantile(0.95)),
                "P99": float(scores.quantile(0.99)),
                "最大值": float(scores.max()),
                "零值占比": float(scores.eq(0).mean()),
                "窗口收益率正值占比": float(group["window_return"].gt(0).mean()),
            }
        )
    return pd.DataFrame.from_records(records)


def build_parameters(score_method: str) -> pd.DataFrame:
    rows = [
        ("得分公式", score_method, SCORE_COLUMNS[score_method]),
        ("因子形成时点", "信号日收盘后", "避免使用同日尚未完成的数据交易"),
        ("ETF买入价", "下一ETF交易日VWAP", "可交易收益主口径"),
        ("ETF卖出价", "持有1至20个ETF交易日后的VWAP", "固定持有期"),
        ("指数收益", "与ETF相同买卖日的指数收盘价", "因子纯度辅助口径"),
        ("主Rank IC范围", SCOPE_ALL, "判断整个截面的排序能力"),
        (
            "局部检验范围",
            "窗口收益率>0、前20%、前10%",
            "先按分数选池，再过滤窗口收益率<=0，不补位",
        ),
        ("最小截面样本", MIN_CROSS_SECTION, "不足时保留记录但不计算IC"),
        ("Q1-Q5方向", "Q1最低分，Q5最高分", "全池等权分组"),
        ("交易成本", TRANSACTION_COST_RATE, "买入和卖出各0.1%"),
        ("Q5-Q1扣费后", "毛收益减4倍单边费率", "多空两端各完成一次买卖"),
        ("显著性", "Newey-West", "滞后阶数=持有期-1"),
        ("滚动窗口", ROLLING_IC_WINDOW, "按有效信号日观测滚动"),
        ("风险中性化", "未做", "ETF轮动先检验原始信号，不混入额外模型"),
    ]
    return pd.DataFrame(rows, columns=["项目", "取值", "说明"])


def add_extreme_return(
    heap: list[tuple[float, int, dict[str, object]]],
    serial: int,
    record: dict[str, object],
    capacity: int,
) -> None:
    item = (abs(float(record["未来收益率"])), serial, record)
    if len(heap) < capacity:
        heapq.heappush(heap, item)
    elif item[0] > heap[0][0]:
        heapq.heapreplace(heap, item)


def research_score_method(
    signals: pd.DataFrame,
    score_method: str,
    mapping: dict[tuple[date, str], CandidateETF],
    schedule: dict[tuple[date, int], tuple[date, date]],
    etf_prices: dict[tuple[date, str], float],
    index_prices: dict[tuple[float, date, str], float],
) -> ResearchFrames:
    method_signals = signals.loc[signals["score_method"].eq(score_method)].copy()
    if method_signals.empty:
        raise ValueError(f"没有找到{score_method}的有效因子记录")

    ic_series: dict[
        tuple[float, int, int, str, str], list[tuple[date, float, int]]
    ] = defaultdict(list)
    ic_attempts: dict[tuple[float, int, int, str, str], int] = defaultdict(int)
    ic_sample_sizes: dict[tuple[float, int, int, str, str], list[int]] = (
        defaultdict(list)
    )
    daily_ic_records: list[dict[str, object]] = []
    coverage: dict[tuple[float, int, int], CoverageStats] = defaultdict(CoverageStats)
    quantile_series: dict[
        tuple[float, int, int, int, str], list[tuple[date, float, int]]
    ] = defaultdict(list)
    long_short_series: dict[
        tuple[float, int, int, str], list[tuple[date, float, int]]
    ] = defaultdict(list)
    top_series: dict[
        tuple[float, int, int, str, str, str], list[tuple[date, float, int]]
    ] = defaultdict(list)

    extreme_capacity = MAX_ANOMALY_ROWS * 2 // 3
    missing_capacity = MAX_ANOMALY_ROWS - extreme_capacity
    extreme_returns: list[tuple[float, int, dict[str, object]]] = []
    missing_samples: list[dict[str, object]] = []
    anomaly_serial = 0

    group_columns = ["threshold", "trend_window", "signal_date"]
    for (threshold, trend_window, signal_date), group in method_signals.groupby(
        group_columns, sort=True
    ):
        threshold = float(threshold)
        trend_window = int(trend_window)
        group = group.sort_values(
            ["score", "index_code"], ascending=[False, True]
        ).reset_index(drop=True)
        scopes = scope_index_codes(group)

        for holding_day in HOLDING_DAYS:
            coverage_key = (threshold, trend_window, holding_day)
            stats = coverage[coverage_key]
            stats.signal_days += 1
            stats.factor_observations += len(group)

            dates = schedule.get((signal_date, holding_day))
            if dates is None:
                for scope_name in scopes:
                    bases = (
                        (RETURN_BASIS_ETF, RETURN_BASIS_INDEX)
                        if scope_name == SCOPE_ALL
                        else (RETURN_BASIS_ETF,)
                    )
                    for return_basis in bases:
                        key = (
                            threshold,
                            trend_window,
                            holding_day,
                            scope_name,
                            return_basis,
                        )
                        ic_attempts[key] += 1
                        ic_sample_sizes[key].append(0)
                for return_basis in (RETURN_BASIS_ETF, RETURN_BASIS_INDEX):
                    daily_ic_records.append(
                        {
                            "信号日": signal_date,
                            "买入日": None,
                            "卖出日": None,
                            "聚类阈值": threshold,
                            "趋势窗口": trend_window,
                            "持有期(交易日)": holding_day,
                            "检验范围": SCOPE_ALL,
                            "收益口径": return_basis,
                            "有效样本数": 0,
                            "RankIC": None,
                            "状态": "无完整持有期",
                        }
                    )
                continue

            stats.scheduled_days += 1
            stats.scheduled_factor_observations += len(group)
            entry_date, exit_date = dates
            assets: list[DailyAssetReturn] = []
            for row in group[
                ["index_code", "score", "window_return"]
            ].itertuples(index=False):
                index_code = str(row.index_code)
                candidate = mapping.get((signal_date, index_code))
                etf_gross_return = None
                etf_net_return = None
                etf_code = candidate.code if candidate is not None else None
                if candidate is not None:
                    stats.mapped_observations += 1
                    entry_vwap = etf_prices.get((entry_date, candidate.code))
                    exit_vwap = etf_prices.get((exit_date, candidate.code))
                    if entry_vwap is not None and exit_vwap is not None:
                        etf_gross_return = exit_vwap / entry_vwap - 1.0
                        etf_net_return = round_trip_net_return(etf_gross_return)
                        stats.etf_return_observations += 1

                entry_close = index_prices.get(
                    (threshold, entry_date, index_code)
                )
                exit_close = index_prices.get((threshold, exit_date, index_code))
                index_return = None
                if entry_close is not None and exit_close is not None:
                    index_return = exit_close / entry_close - 1.0
                    stats.index_return_observations += 1

                asset = DailyAssetReturn(
                    index_code=index_code,
                    etf_code=etf_code,
                    score=float(row.score),
                    window_return=float(row.window_return),
                    etf_gross_return=etf_gross_return,
                    etf_net_return=etf_net_return,
                    index_return=index_return,
                )
                assets.append(asset)

                if holding_day == 1 and etf_gross_return is not None:
                    anomaly_serial += 1
                    add_extreme_return(
                        extreme_returns,
                        anomaly_serial,
                        {
                            "类型": "ETF单日极端收益",
                            "信号日": signal_date,
                            "买入日": entry_date,
                            "卖出日": exit_date,
                            "聚类阈值": threshold,
                            "趋势窗口": trend_window,
                            "对标指数代码": index_code,
                            "ETF代码": etf_code,
                            "因子得分": float(row.score),
                            "未来收益率": etf_gross_return,
                            "说明": "按绝对收益率排序，仅供核查行情、复权和涨跌停",
                        },
                        extreme_capacity,
                    )
                elif (
                    holding_day == 1
                    and len(missing_samples) < missing_capacity
                    and (candidate is None or etf_gross_return is None)
                ):
                    missing_samples.append(
                        {
                            "类型": "ETF映射或价格缺失",
                            "信号日": signal_date,
                            "买入日": entry_date,
                            "卖出日": exit_date,
                            "聚类阈值": threshold,
                            "趋势窗口": trend_window,
                            "对标指数代码": index_code,
                            "ETF代码": etf_code,
                            "因子得分": float(row.score),
                            "未来收益率": None,
                            "说明": (
                                "信号日无可用ETF映射"
                                if candidate is None
                                else "买入日或卖出日缺少有效VWAP"
                            ),
                        }
                    )

            assets_by_code = {asset.index_code: asset for asset in assets}
            for scope_name, index_codes in scopes.items():
                scope_assets = [
                    assets_by_code[index_code]
                    for index_code in index_codes
                    if index_code in assets_by_code
                ]
                bases = (
                    (RETURN_BASIS_ETF, RETURN_BASIS_INDEX)
                    if scope_name == SCOPE_ALL
                    else (RETURN_BASIS_ETF,)
                )
                for return_basis in bases:
                    key = (
                        threshold,
                        trend_window,
                        holding_day,
                        scope_name,
                        return_basis,
                    )
                    ic_attempts[key] += 1
                    valid_pairs = [
                        (asset.score, return_value)
                        for asset in scope_assets
                        for return_value in [
                            asset.etf_gross_return
                            if return_basis == RETURN_BASIS_ETF
                            else asset.index_return
                        ]
                        if return_value is not None
                    ]
                    ic_sample_sizes[key].append(len(valid_pairs))
                    scores = [pair[0] for pair in valid_pairs]
                    future_returns = [pair[1] for pair in valid_pairs]
                    rank_ic, status = rank_ic_status(scores, future_returns)
                    if rank_ic is not None:
                        ic_series[key].append(
                            (signal_date, rank_ic, len(valid_pairs))
                        )
                    if scope_name == SCOPE_ALL:
                        daily_ic_records.append(
                            {
                                "信号日": signal_date,
                                "买入日": entry_date,
                                "卖出日": exit_date,
                                "聚类阈值": threshold,
                                "趋势窗口": trend_window,
                                "持有期(交易日)": holding_day,
                                "检验范围": scope_name,
                                "收益口径": return_basis,
                                "有效样本数": len(valid_pairs),
                                "RankIC": rank_ic,
                                "状态": status,
                            }
                        )

            ranked_scores = pd.Series(
                {asset.index_code: asset.score for asset in assets},
                dtype="float64",
            )
            score_ranks = ranked_scores.rank(method="average", ascending=True)
            quantile_by_code = np.ceil(
                score_ranks / len(score_ranks) * QUANTILE_COUNT
            ).clip(1, QUANTILE_COUNT).astype(int)
            quantile_returns_for_spread: dict[
                str, dict[int, tuple[float, int]]
            ] = {
                "扣费前": {},
                "扣费后": {},
            }
            for quantile in range(1, QUANTILE_COUNT + 1):
                quantile_codes = set(
                    quantile_by_code[quantile_by_code.eq(quantile)].index
                )
                quantile_assets = [
                    asset
                    for asset in assets
                    if asset.index_code in quantile_codes
                    and asset.etf_gross_return is not None
                ]
                if not quantile_assets:
                    continue
                gross_return = float(
                    np.mean([asset.etf_gross_return for asset in quantile_assets])
                )
                net_return = float(
                    np.mean([asset.etf_net_return for asset in quantile_assets])
                )
                for cost_basis, portfolio_return in (
                    ("扣费前", gross_return),
                    ("扣费后", net_return),
                ):
                    quantile_series[
                        (
                            threshold,
                            trend_window,
                            holding_day,
                            quantile,
                            cost_basis,
                        )
                    ].append(
                        (signal_date, portfolio_return, len(quantile_assets))
                    )
                    quantile_returns_for_spread[cost_basis][quantile] = (
                        portfolio_return,
                        len(quantile_assets),
                    )

            gross_quantiles = quantile_returns_for_spread["扣费前"]
            if 1 in gross_quantiles and QUANTILE_COUNT in gross_quantiles:
                gross_spread = (
                    gross_quantiles[QUANTILE_COUNT][0]
                    - gross_quantiles[1][0]
                )
                spread_asset_count = (
                    gross_quantiles[QUANTILE_COUNT][1]
                    + gross_quantiles[1][1]
                )
                for cost_basis, spread in (
                    ("扣费前", gross_spread),
                    (
                        "扣费后",
                        gross_spread - 4.0 * TRANSACTION_COST_RATE,
                    ),
                ):
                    long_short_series[
                        (threshold, trend_window, holding_day, cost_basis)
                    ].append((signal_date, spread, spread_asset_count))

            valid_full_assets = [
                asset for asset in assets if asset.etf_gross_return is not None
            ]
            if valid_full_assets:
                full_returns = {
                    "扣费前": float(
                        np.mean(
                            [asset.etf_gross_return for asset in valid_full_assets]
                        )
                    ),
                    "扣费后": float(
                        np.mean(
                            [asset.etf_net_return for asset in valid_full_assets]
                        )
                    ),
                }
                for scope_name in LOCAL_SCOPE_PERCENTAGES:
                    selected_assets = [
                        assets_by_code[index_code]
                        for index_code in scopes[scope_name]
                        if index_code in assets_by_code
                        and assets_by_code[index_code].etf_gross_return is not None
                    ]
                    if not selected_assets:
                        continue
                    selected_returns = {
                        "扣费前": float(
                            np.mean(
                                [
                                    asset.etf_gross_return
                                    for asset in selected_assets
                                ]
                            )
                        ),
                        "扣费后": float(
                            np.mean(
                                [asset.etf_net_return for asset in selected_assets]
                            )
                        ),
                    }
                    for cost_basis in ("扣费前", "扣费后"):
                        portfolio_return = selected_returns[cost_basis]
                        count = len(selected_assets)
                        top_series[
                            (
                                threshold,
                                trend_window,
                                holding_day,
                                scope_name,
                                cost_basis,
                                "组合收益",
                            )
                        ].append((signal_date, portfolio_return, count))
                        top_series[
                            (
                                threshold,
                                trend_window,
                                holding_day,
                                scope_name,
                                cost_basis,
                                "相对全池超额",
                            )
                        ].append(
                            (
                                signal_date,
                                portfolio_return - full_returns[cost_basis],
                                count,
                            )
                        )

    ic_summary_records: list[dict[str, object]] = []
    all_ic_keys = sorted(ic_attempts)
    for key in all_ic_keys:
        threshold, trend_window, holding_day, scope_name, return_basis = key
        observations = sorted(ic_series.get(key, []))
        metrics = newey_west_statistics(
            (observation[1] for observation in observations),
            lag=holding_day - 1,
        )
        valid_days = int(metrics["样本数"])
        ic_summary_records.append(
            {
                "聚类阈值": threshold,
                "趋势窗口": trend_window,
                "持有期(交易日)": holding_day,
                "检验范围": scope_name,
                "收益口径": return_basis,
                "尝试信号日数": ic_attempts[key],
                "有效IC日数": valid_days,
                "无效IC日数": ic_attempts[key] - valid_days,
                "有效日比例": (
                    valid_days / ic_attempts[key] if ic_attempts[key] else None
                ),
                "日均有效截面数": (
                    float(np.mean(ic_sample_sizes[key]))
                    if ic_sample_sizes[key]
                    else None
                ),
                "RankIC均值": metrics["均值"],
                "RankIC中位数": metrics["中位数"],
                "RankIC标准差": metrics["标准差"],
                "ICIR": metrics["ICIR"],
                "年化ICIR": metrics["年化ICIR"],
                "IC为正比例": metrics["正值比例"],
                "IC大于0.02比例": metrics["大于0.02比例"],
                "NW_t值": metrics["NW_t值"],
                "NW_p值": metrics["NW_p值"],
                "起始信号日": observations[0][0] if observations else None,
                "结束信号日": observations[-1][0] if observations else None,
            }
        )
    ic_summary = pd.DataFrame.from_records(ic_summary_records)

    daily_ic = pd.DataFrame.from_records(daily_ic_records).sort_values(
        ["聚类阈值", "趋势窗口", "收益口径", "持有期(交易日)", "信号日"]
    ).reset_index(drop=True)

    annual_records: list[dict[str, object]] = []
    for key, observations in sorted(ic_series.items()):
        threshold, trend_window, holding_day, scope_name, return_basis = key
        if scope_name != SCOPE_ALL:
            continue
        observations_by_year: dict[int, list[float]] = defaultdict(list)
        for signal_date, rank_ic, _ in observations:
            observations_by_year[signal_date.year].append(rank_ic)
        for year, values in sorted(observations_by_year.items()):
            metrics = newey_west_statistics(values, lag=holding_day - 1)
            annual_records.append(
                {
                    "聚类阈值": threshold,
                    "趋势窗口": trend_window,
                    "持有期(交易日)": holding_day,
                    "收益口径": return_basis,
                    "年份": year,
                    "有效IC日数": metrics["样本数"],
                    "RankIC均值": metrics["均值"],
                    "RankIC中位数": metrics["中位数"],
                    "ICIR": metrics["ICIR"],
                    "IC为正比例": metrics["正值比例"],
                    "NW_t值": metrics["NW_t值"],
                    "NW_p值": metrics["NW_p值"],
                }
            )
    annual_ic = pd.DataFrame.from_records(annual_records)

    rolling_records: list[dict[str, object]] = []
    rolling_min_periods = max(20, ROLLING_IC_WINDOW // 3)
    for key, observations in sorted(ic_series.items()):
        threshold, trend_window, holding_day, scope_name, return_basis = key
        if scope_name != SCOPE_ALL or holding_day not in ROLLING_HOLDING_DAYS:
            continue
        observations = sorted(observations)
        series = pd.Series(
            [item[1] for item in observations],
            index=[item[0] for item in observations],
            dtype="float64",
        )
        rolling_mean = series.rolling(
            ROLLING_IC_WINDOW,
            min_periods=rolling_min_periods,
        ).mean()
        rolling_std = series.rolling(
            ROLLING_IC_WINDOW,
            min_periods=rolling_min_periods,
        ).std(ddof=1)
        for current_date, mean in rolling_mean.dropna().items():
            standard_deviation = rolling_std.loc[current_date]
            rolling_records.append(
                {
                    "信号日": current_date,
                    "聚类阈值": threshold,
                    "趋势窗口": trend_window,
                    "持有期(交易日)": holding_day,
                    "收益口径": return_basis,
                    "滚动窗口": ROLLING_IC_WINDOW,
                    "滚动RankIC均值": float(mean),
                    "滚动ICIR": (
                        float(mean / standard_deviation)
                        if pd.notna(standard_deviation)
                        and standard_deviation > 0
                        else None
                    ),
                }
            )
    rolling_ic = pd.DataFrame.from_records(rolling_records)

    def summarize_return_series(
        source: dict[tuple, list[tuple[date, float, int]]],
        key_names: Sequence[str],
    ) -> pd.DataFrame:
        records: list[dict[str, object]] = []
        for key, observations in sorted(source.items()):
            values = [item[1] for item in observations]
            holding_day = int(key[2])
            metrics = newey_west_statistics(values, lag=holding_day - 1)
            record = dict(zip(key_names, key))
            record.update(
                {
                    "有效信号日数": metrics["样本数"],
                    "日均有效持仓数": float(
                        np.mean([item[2] for item in observations])
                    ),
                    "平均持有期收益": metrics["均值"],
                    "中位数持有期收益": metrics["中位数"],
                    "持有期收益标准差": metrics["标准差"],
                    "正收益比例": metrics["正值比例"],
                    "NW_t值": metrics["NW_t值"],
                    "NW_p值": metrics["NW_p值"],
                    "起始信号日": min(item[0] for item in observations),
                    "结束信号日": max(item[0] for item in observations),
                }
            )
            records.append(record)
        return pd.DataFrame.from_records(records)

    quantile_returns = summarize_return_series(
        quantile_series,
        ["聚类阈值", "趋势窗口", "持有期(交易日)", "分组", "成本口径"],
    )
    if not quantile_returns.empty:
        quantile_returns["分组"] = quantile_returns["分组"].map(
            lambda value: f"Q{int(value)}"
        )

    long_short_returns = summarize_return_series(
        long_short_series,
        ["聚类阈值", "趋势窗口", "持有期(交易日)", "成本口径"],
    )
    if not long_short_returns.empty:
        long_short_returns.insert(3, "组合", "Q5-Q1")

    top_portfolios = summarize_return_series(
        top_series,
        [
            "聚类阈值",
            "趋势窗口",
            "持有期(交易日)",
            "候选范围",
            "成本口径",
            "统计项目",
        ],
    )

    quality_records: list[dict[str, object]] = []
    for key, stats in sorted(coverage.items()):
        threshold, trend_window, holding_day = key
        quality_records.append(
            {
                "聚类阈值": threshold,
                "趋势窗口": trend_window,
                "持有期(交易日)": holding_day,
                "信号日数": stats.signal_days,
                "有完整持有期日数": stats.scheduled_days,
                "完整持有期覆盖率": (
                    stats.scheduled_days / stats.signal_days
                    if stats.signal_days
                    else None
                ),
                "因子观测数": stats.factor_observations,
                "完整持有期因子观测数": stats.scheduled_factor_observations,
                "成功映射ETF数": stats.mapped_observations,
                "ETF映射覆盖率": (
                    stats.mapped_observations
                    / stats.scheduled_factor_observations
                    if stats.scheduled_factor_observations
                    else None
                ),
                "有效ETF收益数": stats.etf_return_observations,
                "ETF收益覆盖率": (
                    stats.etf_return_observations
                    / stats.scheduled_factor_observations
                    if stats.scheduled_factor_observations
                    else None
                ),
                "有效指数收益数": stats.index_return_observations,
                "指数收益覆盖率": (
                    stats.index_return_observations
                    / stats.scheduled_factor_observations
                    if stats.scheduled_factor_observations
                    else None
                ),
            }
        )
    data_quality = pd.DataFrame.from_records(quality_records)

    robustness_records: list[dict[str, object]] = []
    ic_robustness_columns = [
        "聚类阈值",
        "趋势窗口",
        "持有期(交易日)",
        "RankIC均值",
        "有效IC日数",
        "NW_t值",
        "NW_p值",
    ]
    for row in ic_summary.loc[
        ic_summary["检验范围"].eq(SCOPE_ALL)
        & ic_summary["收益口径"].eq(RETURN_BASIS_ETF),
        ic_robustness_columns,
    ].itertuples(index=False, name=None):
        (
            threshold,
            trend_window,
            holding_day,
            mean_value,
            valid_days,
            t_value,
            p_value,
        ) = row
        robustness_records.append(
            {
                "检验项": "全池ETF RankIC",
                "聚类阈值": threshold,
                "趋势窗口": trend_window,
                "持有期(交易日)": holding_day,
                "候选范围": SCOPE_ALL,
                "成本口径": "不适用",
                "均值": mean_value,
                "有效日数": valid_days,
                "NW_t值": t_value,
                "NW_p值": p_value,
            }
        )
    if not long_short_returns.empty:
        return_columns = [
            "聚类阈值",
            "趋势窗口",
            "持有期(交易日)",
            "平均持有期收益",
            "有效信号日数",
            "NW_t值",
            "NW_p值",
        ]
        for row in long_short_returns.loc[
            long_short_returns["成本口径"].eq("扣费后"), return_columns
        ].itertuples(index=False, name=None):
            (
                threshold,
                trend_window,
                holding_day,
                mean_value,
                valid_days,
                t_value,
                p_value,
            ) = row
            robustness_records.append(
                {
                    "检验项": "Q5-Q1收益",
                    "聚类阈值": threshold,
                    "趋势窗口": trend_window,
                    "持有期(交易日)": holding_day,
                    "候选范围": "Q5-Q1",
                    "成本口径": "扣费后",
                    "均值": mean_value,
                    "有效日数": valid_days,
                    "NW_t值": t_value,
                    "NW_p值": p_value,
                }
            )
    if not top_portfolios.empty:
        top_excess = top_portfolios.loc[
            top_portfolios["成本口径"].eq("扣费后")
            & top_portfolios["统计项目"].eq("相对全池超额")
        ]
        top_columns = [
            "聚类阈值",
            "趋势窗口",
            "持有期(交易日)",
            "候选范围",
            "平均持有期收益",
            "有效信号日数",
            "NW_t值",
            "NW_p值",
        ]
        for row in top_excess[top_columns].itertuples(index=False, name=None):
            (
                threshold,
                trend_window,
                holding_day,
                scope_name,
                mean_value,
                valid_days,
                t_value,
                p_value,
            ) = row
            robustness_records.append(
                {
                    "检验项": "Top组合相对全池超额",
                    "聚类阈值": threshold,
                    "趋势窗口": trend_window,
                    "持有期(交易日)": holding_day,
                    "候选范围": scope_name,
                    "成本口径": "扣费后",
                    "均值": mean_value,
                    "有效日数": valid_days,
                    "NW_t值": t_value,
                    "NW_p值": p_value,
                }
            )
    robustness = pd.DataFrame.from_records(robustness_records)

    anomaly_records = [item[2] for item in sorted(extreme_returns, reverse=True)]
    anomaly_records.extend(missing_samples)
    anomalies = pd.DataFrame.from_records(anomaly_records)

    return ResearchFrames(
        parameters=build_parameters(score_method),
        factor_distribution=calculate_factor_distribution(method_signals),
        data_quality=data_quality,
        ic_summary=ic_summary,
        daily_ic=daily_ic,
        annual_ic=annual_ic,
        rolling_ic=rolling_ic,
        quantile_returns=quantile_returns,
        long_short_returns=long_short_returns,
        top_portfolios=top_portfolios,
        robustness=robustness,
        anomalies=anomalies,
    )


def set_sheet_style(worksheet, frame: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_number, column_name in enumerate(frame.columns, start=1):
        values = frame[column_name].fillna("").astype(str)
        maximum_length = values.map(len).max() if not values.empty else 0
        width = min(max(len(str(column_name)), maximum_length) + 2, 24)
        worksheet.column_dimensions[get_column_letter(column_number)].width = width

        if any(
            keyword in str(column_name)
            for keyword in ("覆盖率", "比例", "占比", "收益")
        ):
            for cell in worksheet[get_column_letter(column_number)][1:]:
                cell.number_format = "0.0000%"


def write_frame(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    if len(frame) > EXCEL_MAX_DATA_ROWS:
        raise ValueError(
            f"{sheet_name}共有{len(frame):,}行，超过Excel单表上限，"
            "请缩小参数范围或拆分输出"
        )
    frame.to_excel(writer, sheet_name=sheet_name, index=False)
    set_sheet_style(writer.sheets[sheet_name], frame)


def configure_chinese_plotting() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Arial Unicode MS",
        "PingFang SC",
        "Heiti SC",
        "SimHei",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def save_ic_decay_chart(frames: ResearchFrames, output_file: Path) -> None:
    if frames.ic_summary.empty:
        return
    data = frames.ic_summary.loc[
        frames.ic_summary["检验范围"].eq(SCOPE_ALL)
        & frames.ic_summary["收益口径"].eq(RETURN_BASIS_ETF)
    ]
    if data.empty:
        return
    figure, axis = plt.subplots(figsize=(12, 7))
    for (threshold, trend_window), group in data.groupby(
        ["聚类阈值", "趋势窗口"], sort=True
    ):
        axis.plot(
            group["持有期(交易日)"],
            group["RankIC均值"],
            marker="o",
            markersize=3,
            linewidth=1.3,
            label=f"阈值{threshold:g}/窗口{trend_window}",
        )
    axis.axhline(0, color="black", linewidth=0.8)
    axis.set(title="全池ETF RankIC衰减", xlabel="持有期（交易日）", ylabel="RankIC均值")
    axis.grid(alpha=0.25)
    axis.legend(ncol=3, fontsize=8)
    figure.tight_layout()
    figure.savefig(output_file, dpi=180)
    plt.close(figure)


def save_quantile_chart(frames: ResearchFrames, output_file: Path) -> None:
    if frames.quantile_returns.empty:
        return
    data = frames.quantile_returns.loc[
        frames.quantile_returns["成本口径"].eq("扣费后")
    ]
    if data.empty:
        return
    configurations = list(
        data[["聚类阈值", "趋势窗口"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    )
    column_count = min(3, len(configurations))
    row_count = math.ceil(len(configurations) / column_count)
    figure, axes = plt.subplots(
        row_count,
        column_count,
        figsize=(5 * column_count, 3.6 * row_count),
        squeeze=False,
        sharex=True,
    )
    for axis, (threshold, trend_window) in zip(axes.flat, configurations):
        group = data.loc[
            data["聚类阈值"].eq(threshold)
            & data["趋势窗口"].eq(trend_window)
        ]
        for quantile, quantile_group in group.groupby("分组", sort=True):
            axis.plot(
                quantile_group["持有期(交易日)"],
                quantile_group["平均持有期收益"],
                marker="o",
                markersize=2.5,
                linewidth=1,
                label=quantile,
            )
        axis.axhline(0, color="black", linewidth=0.7)
        axis.set_title(f"阈值{threshold:g} / 窗口{trend_window}")
        axis.yaxis.set_major_formatter(PercentFormatter(1.0))
        axis.grid(alpha=0.2)
    for axis in axes.flat[len(configurations):]:
        axis.set_visible(False)
    handles, labels = axes.flat[0].get_legend_handles_labels()
    figure.legend(handles, labels, loc="upper center", ncol=QUANTILE_COUNT)
    figure.suptitle("Q1-Q5扣费后平均持有期收益", y=1.01)
    figure.tight_layout()
    figure.savefig(output_file, dpi=180, bbox_inches="tight")
    plt.close(figure)


def save_rolling_ic_chart(frames: ResearchFrames, output_file: Path) -> None:
    if frames.rolling_ic.empty:
        return
    data = frames.rolling_ic.loc[
        frames.rolling_ic["收益口径"].eq(RETURN_BASIS_ETF)
        & frames.rolling_ic["持有期(交易日)"].eq(10)
    ]
    if data.empty:
        return
    figure, axis = plt.subplots(figsize=(13, 7))
    for (threshold, trend_window), group in data.groupby(
        ["聚类阈值", "趋势窗口"], sort=True
    ):
        axis.plot(
            pd.to_datetime(group["信号日"]),
            group["滚动RankIC均值"],
            linewidth=1,
            label=f"阈值{threshold:g}/窗口{trend_window}",
        )
    axis.axhline(0, color="black", linewidth=0.8)
    axis.set(title="60期滚动RankIC（持有10日）", xlabel="信号日", ylabel="滚动RankIC均值")
    axis.grid(alpha=0.2)
    axis.legend(ncol=3, fontsize=8)
    figure.tight_layout()
    figure.savefig(output_file, dpi=180)
    plt.close(figure)


def save_parameter_heatmap(frames: ResearchFrames, output_file: Path) -> None:
    if frames.ic_summary.empty:
        return
    data = frames.ic_summary.loc[
        frames.ic_summary["检验范围"].eq(SCOPE_ALL)
        & frames.ic_summary["收益口径"].eq(RETURN_BASIS_ETF)
        & frames.ic_summary["持有期(交易日)"].eq(10)
    ]
    if data.empty:
        return
    matrix = data.pivot(
        index="聚类阈值",
        columns="趋势窗口",
        values="RankIC均值",
    ).sort_index().sort_index(axis=1)
    figure, axis = plt.subplots(figsize=(7, 5))
    image = axis.imshow(matrix.to_numpy(), cmap="RdYlGn", aspect="auto")
    axis.set_xticks(range(len(matrix.columns)), labels=matrix.columns)
    axis.set_yticks(
        range(len(matrix.index)), labels=[f"{value:g}" for value in matrix.index]
    )
    axis.set(xlabel="趋势窗口", ylabel="聚类阈值", title="持有10日全池ETF RankIC参数热力图")
    for row_number in range(len(matrix.index)):
        for column_number in range(len(matrix.columns)):
            value = matrix.iloc[row_number, column_number]
            if pd.notna(value):
                axis.text(
                    column_number,
                    row_number,
                    f"{value:.3f}",
                    ha="center",
                    va="center",
                    fontsize=9,
                )
    figure.colorbar(image, ax=axis, label="RankIC均值")
    figure.tight_layout()
    figure.savefig(output_file, dpi=180)
    plt.close(figure)


def write_output(score_method: str, frames: ResearchFrames) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_FILES[score_method]
    sheets = {
        "参数与口径": frames.parameters,
        "因子分布": frames.factor_distribution,
        "数据质量": frames.data_quality,
        "汇总": frames.ic_summary,
        "逐日RankIC": frames.daily_ic,
        "年度IC": frames.annual_ic,
        "滚动IC": frames.rolling_ic,
        "分组收益": frames.quantile_returns,
        "多空收益": frames.long_short_returns,
        "Top组合": frames.top_portfolios,
        "稳健性": frames.robustness,
        "异常样本": frames.anomalies,
    }
    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
        date_format="yyyy-mm-dd",
    ) as writer:
        for sheet_name, frame in sheets.items():
            write_frame(writer, sheet_name, frame)

    configure_chinese_plotting()
    chart_prefix = output_file.with_suffix("")
    save_ic_decay_chart(frames, Path(f"{chart_prefix}_IC衰减.png"))
    save_quantile_chart(frames, Path(f"{chart_prefix}_Q1-Q5.png"))
    save_rolling_ic_chart(frames, Path(f"{chart_prefix}_滚动IC.png"))
    save_parameter_heatmap(frames, Path(f"{chart_prefix}_参数热力图.png"))


def main() -> None:
    print("读取趋势因子……")
    signals = load_factor_signals()
    print("扫描ETF数据并建立信号日映射……")
    mapping, trading_dates = scan_etf_mapping(signals)
    schedule = build_trade_schedule(set(signals["signal_date"]), trading_dates)
    required_prices = collect_required_prices(signals, mapping, schedule)
    print("读取计算未来收益所需的VWAP……")
    etf_prices = load_required_vwap(required_prices)
    print("读取对标指数收盘价……")
    index_prices = load_index_closes(signals)
    for score_method in SCORE_COLUMNS:
        print(f"检验{score_method}……")
        frames = research_score_method(
            signals,
            score_method,
            mapping,
            schedule,
            etf_prices,
            index_prices,
        )
        write_output(score_method, frames)
    print(f"完成：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
